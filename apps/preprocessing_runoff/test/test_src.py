import datetime as dt
import os
import sys

import pandas as pd
import pytest

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
import src

# Helper to get absolute paths to test files
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_FILES_DIR = os.path.join(TEST_DIR, "test_files")


def test_get_runoff_data_no_data_available():
    os.environ["ieasyforecast_daily_discharge_path"] = os.path.join(TEST_FILES_DIR, "test_config")

    output = src.get_runoff_data()
    print("Output: ")
    print(output)

    os.environ.pop("ieasyforecast_daily_discharge_path")


def test_read_runoff_data_from_multiple_rivers_xlsx():
    filename = os.path.join(TEST_FILES_DIR, "test_runoff_file.xlsx")
    expected_output = pd.DataFrame(
        {
            "date": [
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
            ],
            "discharge": [2.3, 2.4, 2.5, 2.6, 2.7, 4.3, 4.4, 4.5, 4.6, 4.7],
            "name": [
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
            ],
            "code": [17123, 17123, 17123, 17123, 17123, 17456, 17456, 17456, 17456, 17456],
        }
    ).reset_index(drop=True)
    expected_output["date"] = pd.to_datetime(expected_output["date"]).dt.normalize()

    output = src.read_runoff_data_from_multiple_rivers_xlsx(
        filename, code_list=["17123", "17456"]
    ).reset_index(drop=True)

    assert output.equals(expected_output)


def test_read_runoff_data_from_multiple_rivers_no_code():
    filename = os.path.join(TEST_FILES_DIR, "files_with_errors", "test_runoff_file_no_code.xlsx")

    expected_output = pd.DataFrame(
        {
            "date": ["2000-01-01", "2000-01-02", "2000-01-03", "2000-01-04", "2000-01-05"],
            "discharge": [2.3, 2.4, 2.5, 2.6, 2.7],
            "name": [
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
            ],
            "code": [17123, 17123, 17123, 17123, 17123],
        }
    ).reset_index(drop=True)
    expected_output["date"] = pd.to_datetime(expected_output["date"]).dt.normalize()

    output = src.read_runoff_data_from_multiple_rivers_xlsx(
        filename, code_list=["17123"]
    ).reset_index(drop=True)

    # assert if all values in column discharge are NaN
    assert output.equals(expected_output)


def test_read_runoff_data_from_multiple_rivers_without_data_in_xls():
    filename = os.path.join(TEST_FILES_DIR, "files_with_errors", "test_runoff_file_no_data.xlsx")

    output = src.read_runoff_data_from_multiple_rivers_xlsx(
        filename, code_list=["17123"]
    ).reset_index(drop=True)

    # assert if all values in column discharge are NaN
    assert output["discharge"].isna().all()


def test_read_runoff_data_from_multiple_rivers_no_file():
    filename = os.path.join(TEST_FILES_DIR, "files_with_errors", "this_file_does_not_exist.xlsx")

    # Assert FileNotFoundError is raised
    with pytest.raises(FileNotFoundError):
        src.read_runoff_data_from_multiple_rivers_xlsx(filename, code_list=[123])


def test_read_runoff_data_from_multiple_rivers_no_station_header():
    filename = os.path.join(
        TEST_FILES_DIR, "files_with_errors", "test_runoff_file_no_station_header.xlsx"
    )

    with pytest.raises(ValueError):
        src.read_runoff_data_from_multiple_rivers_xlsx(filename, code_list=[123])


def test_read_all_runoff_data_from_excel():
    expected_output = pd.DataFrame(
        {
            "date": [
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
                "2001-01-01",
                "2001-01-02",
                "2001-01-03",
                "2001-01-04",
                "2001-01-05",
            ],
            "discharge": [
                2.3,
                2.4,
                2.5,
                2.6,
                2.7,
                4.3,
                4.4,
                4.5,
                4.6,
                4.7,
                2.3,
                2.4,
                2.5,
                2.6,
                2.7,
                4.3,
                4.4,
                4.5,
                4.6,
                4.7,
            ],
            "name": [
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "s. n. wi - spec ch",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
                "other r. - hi",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ],
            "code": [
                17123,
                17123,
                17123,
                17123,
                17123,
                17456,
                17456,
                17456,
                17456,
                17456,
                12345,
                12345,
                12345,
                12345,
                12345,
                12345,
                12345,
                12345,
                12345,
                12345,
            ],
        }
    ).reset_index(drop=True)
    expected_output["date"] = pd.to_datetime(expected_output["date"]).dt.normalize()

    os.environ["ieasyforecast_daily_discharge_path"] = TEST_FILES_DIR

    output = src.read_all_runoff_data_from_excel(code_list=["17123", "17456", "12345"]).reset_index(
        drop=True
    )

    os.environ.pop("ieasyforecast_daily_discharge_path")

    assert output.equals(expected_output)


def test_write_data_to_csv():
    runoff_data = pd.DataFrame(
        {
            "date": ["2000-01-01", "2000-01-02", "2000-01-03", "2000-01-04", "2000-01-05"],
            "discharge": [2.3, 2.4, 2.5, 2.6, 2.7],
            "name": ["a", "a", "a", "a", "a"],
            "code": [1, 1, 1, 1, 1],
        }
    )

    # Define environment variables
    os.environ["ieasyforecast_intermediate_data_path"] = TEST_FILES_DIR
    os.environ["ieasyforecast_daily_discharge_file"] = "test_runoff_file.csv"

    # Write the output file
    src.write_daily_time_series_data_to_csv(runoff_data)

    # Read the output file
    output_file = os.path.join(TEST_FILES_DIR, "test_runoff_file.csv")
    output = pd.read_csv(output_file)

    # The data in columns date, discharge and code should be the same
    assert output["date"].equals(runoff_data["date"])
    assert output["discharge"].equals(runoff_data["discharge"])
    assert output["code"].equals(runoff_data["code"])

    # Clean up the environment variables
    os.environ.pop("ieasyforecast_intermediate_data_path")
    os.environ.pop("ieasyforecast_daily_discharge_file")

    # Remove the output file
    os.remove(output_file)


def test_filter_roughly_for_outliers_no_outliers():
    # Create a DataFrame with no outliers
    df = pd.DataFrame(
        {
            "Date": [
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
            ],
            "Code": ["A", "A", "A", "B", "B", "B"],
            "Q_m3s": [1.0, 2.0, 3.0, 4.0, 5.0, 6.0],
        }
    )
    # Convert df['Date'] to datetime
    df["Date"] = pd.to_datetime(df["Date"])

    # Apply the function
    result = src.filter_roughly_for_outliers(df, "Code", "Q_m3s", "Date")

    # Drop index
    result = result.reset_index(drop=True)

    # Check that all original data is preserved (values should match for original dates)
    for code in ["A", "B"]:
        for date in ["2000-01-01", "2000-01-02", "2000-01-03"]:
            orig_val = df[(df["Code"] == code) & (df["Date"] == date)]["Q_m3s"].values[0]
            result_val = result[(result["Code"] == code) & (result["Date"] == date)][
                "Q_m3s"
            ].values[0]
            assert orig_val == result_val, f"Value mismatch for {code} on {date}"

    # Verify no NaN values were introduced for original dates
    for code in ["A", "B"]:
        code_data = result[result["Code"] == code]
        original_dates = pd.to_datetime(["2000-01-01", "2000-01-02", "2000-01-03"])
        for date in original_dates:
            val = code_data[code_data["Date"] == date]["Q_m3s"].values[0]
            assert not pd.isna(val), f"Unexpected NaN for {code} on {date}"


def test_filter_roughly_for_outliers_with_outliers():
    # Create a DataFrame with an outlier
    df = pd.DataFrame(
        {
            "Date": [
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-01",
                "2000-01-02",
                "2000-01-03",
                "2000-01-04",
                "2000-01-05",
                "2000-01-06",
                "2000-01-04",
                "2000-01-05",
                "2000-01-06",
                "2000-01-07",
                "2000-01-08",
                "2000-01-09",
                "2000-01-07",
                "2000-01-08",
                "2000-01-09",
                "2000-01-10",
                "2000-01-11",
                "2000-01-12",
                "2000-01-10",
                "2000-01-11",
                "2000-01-12",
                "2000-01-13",
                "2000-01-14",
                "2000-01-15",
                "2000-01-13",
                "2000-01-14",
                "2000-01-15",
                "2000-01-16",
                "2000-01-17",
                "2000-01-18",
                "2000-01-16",
                "2000-01-17",
                "2000-01-18",
                "2000-01-19",
                "2000-01-20",
                "2000-01-21",
                "2000-01-19",
                "2000-01-20",
                "2000-01-21",
                "2000-01-22",
                "2000-01-23",
                "2000-01-24",
                "2000-01-22",
                "2000-01-23",
                "2000-01-24",
                "2000-01-25",
                "2000-01-26",
                "2000-01-27",
                "2000-01-25",
                "2000-01-26",
                "2000-01-27",
                "2000-01-28",
                "2000-01-29",
                "2000-01-30",
                "2000-01-28",
                "2000-01-29",
                "2000-01-30",
                "2000-01-31",
                "2000-02-01",
                "2000-02-02",
                "2000-01-31",
                "2000-02-01",
                "2000-02-02",
            ],
            "Category": [
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
                "A",
                "A",
                "A",
                "B",
                "B",
                "B",
            ],
            "Values": [
                1.01,
                2.01,
                3.01,
                4.0,
                5.0,
                6.0,
                1.02,
                2.02,
                3.02,
                4.0,
                5.0,
                6.0,
                1.03,
                2.03,
                3.03,
                4.0,
                5.0,
                6.0,
                1.04,
                2.04,
                3.04,
                4.0,
                5.0,
                6.0,
                1.05,
                2.05,
                3.05,
                4.0,
                5.0,
                6.0,
                1.06,
                2.06,
                18.0,
                4.0,
                5.0,
                6.0,
                1.07,
                2.07,
                3.07,
                4.0,
                5.0,
                6.0,
                1.08,
                2.08,
                3.08,
                4.0,
                5.0,
                6.0,
                1.09,
                2.09,
                3.09,
                4.0,
                5.0,
                6.0,
                1.10,
                2.10,
                3.10,
                4.0,
                5.0,
                6.0,
                1.11,
                2.11,
                3.11,
                4.0,
                5.0,
                6.0,
            ],
        }
    )
    # Convert df['Date'] to datetime
    df["Date"] = pd.to_datetime(df["Date"])

    # Apply the function
    result = src.filter_roughly_for_outliers(df, "Category", "Values", "Date")
    # Print value on January 18th for category A
    new = result[(result["Category"] == "A") & (result["Date"] == "2000-01-18")]
    old = df[(df["Category"] == "A") & (df["Date"] == "2000-01-18")]

    # Check that the outlier has been replaced with NaN
    # There should be exactly one NaN value in the DataFrame column Q_m3s
    # print(result[result['Values']==100.0])
    # print(result['Values'].isna().sum())
    assert result["Values"].isna().sum() == 0
    # Assert that the outlier has been replaced with the linear interpolation
    assert new["Values"].values[0] != old["Values"].values[0]


class TestFromDailyTimeSeriestoHydrograph:
    """Test class for the from_daily_time_series_to_hydrograph function."""

    def test_leap_year_handling(self):
        """Test proper handling of leap years in hydrograph generation."""
        # Create a DataFrame spanning multiple years, including a leap year
        dates = []
        values = []

        # Create test data with dates from 2019-2021 (2020 is a leap year)
        for year in [2019, 2020, 2021]:
            # Create full year of data
            year_dates = pd.date_range(start=f"{year}-01-01", end=f"{year}-12-31")
            dates.extend(year_dates)

            # Add some test values (just using day of year as the value)
            values.extend([date.dayofyear for date in year_dates])

        # Create DataFrame
        df = pd.DataFrame(
            {"date": dates, "discharge": values, "code": "15194", "name": "Test Site"}
        )

        # Run the function
        result = src.from_daily_time_series_to_hydrograph(df)

        # Check for leap year handling
        # We should have day_of_year values 1-365 (no 366 even though 2020 is a leap year)
        assert set(result["day_of_year"].unique()) == set(range(1, 366))

        # The dates in the result should be in the current year
        current_year = dt.date.today().year
        assert all(d.year == current_year for d in result["date"])

        # Verify that date sequence is continuous (no gaps)
        sorted_result = result.sort_values("date")
        date_diffs = sorted_result["date"].diff().iloc[1:].dt.days
        assert date_diffs.max() == 1
        assert date_diffs.min() == 1

    def test_statistics_calculation(self):
        """Test that statistics are correctly calculated for historical data."""
        # Get current year for testing
        current_year = dt.date.today().year
        last_year = current_year - 1

        # Create 5 years of data for day 1-3 of January with known patterns
        dates = []
        values = []

        for year in range(current_year - 4, current_year + 1):
            for day in range(1, 4):
                dates.append(dt.datetime(year, 1, day))
                if year == current_year:
                    values.append(day * 10)  # Current year values: 10, 20, 30
                elif year == last_year:
                    values.append(day * 5)  # Last year values: 5, 10, 15
                else:
                    values.append(day)  # Earlier years values: 1, 2, 3

        # Create DataFrame
        df = pd.DataFrame(
            {"date": dates, "discharge": values, "code": "15194", "name": "Test Site"}
        )

        # Run the function
        result = src.from_daily_time_series_to_hydrograph(df)

        # Check statistics for each day
        for day in range(1, 4):
            day_result = result[result["date"].dt.day == day].iloc[0]

            # Check count is correct (5 years of data)
            assert day_result["count"] == 5

            # Check mean (3 early years with value=day, last year with 5*day, current year with 10*day)
            expected_mean = (day * 3 + day * 5 + day * 10) / 5
            assert abs(day_result["mean"] - expected_mean) < 0.0001

            # Check percentiles
            assert day_result["min"] == day  # Minimum is just the day value
            assert day_result["max"] == day * 10  # Maximum is current year value

            # Check current and previous year values
            assert day_result[str(current_year)] == day * 10
            assert day_result[str(last_year)] == day * 5

    def test_multiple_sites(self):
        """Test processing of multiple sites within the same dataset."""
        # Create test data for two sites
        dates = pd.date_range(start="2021-01-01", periods=10, freq="D")

        data = []
        for site_code in ["15194", "15212"]:
            for date in dates:
                # Different pattern for each site
                if site_code == "15194":
                    value = date.day
                else:
                    value = date.day * 2

                data.append(
                    {
                        "date": date,
                        "discharge": value,
                        "code": site_code,
                        "name": f"Test Site {site_code}",
                    }
                )

        df = pd.DataFrame(data)

        # Run the function
        result = src.from_daily_time_series_to_hydrograph(df)

        # Verify each site is processed separately
        site_groups = result.groupby("code")
        assert len(site_groups) == 2

        # Check each site has the correct data
        site1_data = site_groups.get_group("15194")
        site2_data = site_groups.get_group("15212")

        # Both sites should have same number of days
        assert len(site1_data) == len(dates)
        assert len(site2_data) == len(dates)

        # Check that means reflect the different patterns
        for day in range(1, 11):
            site1_day = site1_data[site1_data["date"].dt.day == day]
            site2_day = site2_data[site2_data["date"].dt.day == day]

            if not site1_day.empty and not site2_day.empty:
                assert site1_day["mean"].iloc[0] == day
                assert site2_day["mean"].iloc[0] == day * 2


class TestMergeWithUpdate:
    """Tests for the _merge_with_update helper function."""

    def test_merge_empty_existing_data(self):
        """Test merging when existing data is empty."""
        existing = pd.DataFrame(columns=["code", "date", "discharge"])
        new_data = pd.DataFrame(
            {
                "code": ["A", "A"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-02"]),
                "discharge": [10.0, 20.0],
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        assert len(result) == 2
        assert result["discharge"].tolist() == [10.0, 20.0]

    def test_merge_empty_new_data(self):
        """Test merging when new data is empty."""
        existing = pd.DataFrame(
            {
                "code": ["A", "A"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-02"]),
                "discharge": [10.0, 20.0],
            }
        )
        new_data = pd.DataFrame(columns=["code", "date", "discharge"])

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        assert len(result) == 2
        assert result["discharge"].tolist() == [10.0, 20.0]

    def test_merge_updates_existing_values(self):
        """Test that existing values are updated with new values."""
        existing = pd.DataFrame(
            {
                "code": ["A", "A", "B"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-02", "2024-01-01"]),
                "discharge": [10.0, 20.0, 30.0],
            }
        )
        new_data = pd.DataFrame(
            {
                "code": ["A"],
                "date": pd.to_datetime(["2024-01-02"]),
                "discharge": [25.0],  # Updated value
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        # Find the updated row
        updated_row = result[
            (result["code"] == "A") & (result["date"] == pd.Timestamp("2024-01-02"))
        ]
        assert updated_row["discharge"].values[0] == 25.0

    def test_merge_adds_new_rows(self):
        """Test that new rows are added to the result."""
        existing = pd.DataFrame(
            {"code": ["A"], "date": pd.to_datetime(["2024-01-01"]), "discharge": [10.0]}
        )
        new_data = pd.DataFrame(
            {
                "code": ["A", "B"],
                "date": pd.to_datetime(["2024-01-02", "2024-01-01"]),
                "discharge": [20.0, 30.0],
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        assert len(result) == 3  # 1 original + 2 new


class TestLoadCachedData:
    """Tests for the _load_cached_data helper function."""

    def test_loads_existing_csv(self, tmp_path):
        """Test loading data from an existing CSV file."""
        # Create a test CSV file
        test_data = pd.DataFrame(
            {
                "code": ["15194", "15194"],
                "date": ["2024-01-01", "2024-01-02"],
                "discharge": [10.0, 20.0],
            }
        )
        csv_path = tmp_path / "daily_discharge.csv"
        test_data.to_csv(csv_path, index=False)

        # Set environment variables
        original_path = os.environ.get("ieasyforecast_intermediate_data_path")
        original_file = os.environ.get("ieasyforecast_daily_discharge_file")

        try:
            os.environ["ieasyforecast_intermediate_data_path"] = str(tmp_path)
            os.environ["ieasyforecast_daily_discharge_file"] = "daily_discharge.csv"

            result = src._load_cached_data(
                date_col="date",
                discharge_col="discharge",
                name_col="name",
                code_col="code",
                code_list=["15194"],
            )

            assert len(result) == 2
            assert result["code"].tolist() == ["15194", "15194"]
        finally:
            # Restore environment
            if original_path is None:
                os.environ.pop("ieasyforecast_intermediate_data_path", None)
            else:
                os.environ["ieasyforecast_intermediate_data_path"] = original_path
            if original_file is None:
                os.environ.pop("ieasyforecast_daily_discharge_file", None)
            else:
                os.environ["ieasyforecast_daily_discharge_file"] = original_file


class TestMaintenanceModeGapFilling:
    """Tests for maintenance mode gap filling behavior.

    These tests verify that the data merging logic correctly fills gaps
    in cached data. The full integration of get_runoff_data_for_sites_HF
    requires SDK and environment setup, so we test the underlying
    merge functionality directly.
    """

    def test_merge_with_update_fills_missing_rows(self):
        """Test that _merge_with_update correctly fills missing rows.

        This tests the underlying merge mechanism used to fill gaps
        from database data.
        """
        # Existing data with gaps (missing Jan 3-4)
        existing = pd.DataFrame(
            {
                "code": ["15194", "15194", "15194"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-02", "2024-01-05"]),
                "discharge": [10.0, 20.0, 50.0],
            }
        )

        # New data from DB that has the missing dates
        new_data = pd.DataFrame(
            {
                "code": ["15194", "15194"],
                "date": pd.to_datetime(["2024-01-03", "2024-01-04"]),
                "discharge": [30.0, 40.0],
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        # Should have all 5 rows now
        assert len(result) == 5, f"Expected 5 rows, got {len(result)}"

        # Verify the previously missing rows are now present
        jan3 = result[(result["code"] == "15194") & (result["date"] == pd.Timestamp("2024-01-03"))]
        jan4 = result[(result["code"] == "15194") & (result["date"] == pd.Timestamp("2024-01-04"))]

        assert len(jan3) == 1, "Jan 3 data should be present"
        assert len(jan4) == 1, "Jan 4 data should be present"
        assert jan3["discharge"].values[0] == 30.0
        assert jan4["discharge"].values[0] == 40.0

    def test_merge_with_update_fills_multiple_site_gaps(self):
        """Test gap filling works correctly for multiple sites."""
        # Existing data with gaps for multiple sites
        existing = pd.DataFrame(
            {
                "code": ["15194", "15194", "16059", "16059"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-03", "2024-01-01", "2024-01-03"]),
                "discharge": [10.0, 30.0, 100.0, 300.0],
            }
        )

        # New data fills gaps for both sites
        new_data = pd.DataFrame(
            {
                "code": ["15194", "16059"],
                "date": pd.to_datetime(["2024-01-02", "2024-01-02"]),
                "discharge": [20.0, 200.0],
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        # Should have 6 rows total (3 per site)
        assert len(result) == 6, f"Expected 6 rows, got {len(result)}"

        # Verify gaps are filled for site 15194
        site_15194 = result[result["code"] == "15194"].sort_values("date")
        assert len(site_15194) == 3
        assert site_15194["discharge"].tolist() == [10.0, 20.0, 30.0]

        # Verify gaps are filled for site 16059
        site_16059 = result[result["code"] == "16059"].sort_values("date")
        assert len(site_16059) == 3
        assert site_16059["discharge"].tolist() == [100.0, 200.0, 300.0]

    def test_merge_updates_and_adds_simultaneously(self):
        """Test that merge can update existing values AND add new rows."""
        existing = pd.DataFrame(
            {
                "code": ["15194", "15194"],
                "date": pd.to_datetime(["2024-01-01", "2024-01-02"]),
                "discharge": [10.0, 20.0],  # Jan 2 will be updated
            }
        )

        new_data = pd.DataFrame(
            {
                "code": ["15194", "15194"],
                "date": pd.to_datetime(["2024-01-02", "2024-01-03"]),
                "discharge": [25.0, 30.0],  # Updated value for Jan 2, new for Jan 3
            }
        )

        result = src._merge_with_update(existing, new_data, "code", "date", "discharge")

        assert len(result) == 3, f"Expected 3 rows, got {len(result)}"

        # Check Jan 2 was updated
        jan2 = result[result["date"] == pd.Timestamp("2024-01-02")]
        assert jan2["discharge"].values[0] == 25.0, "Jan 2 should be updated to 25.0"

        # Check Jan 3 was added
        jan3 = result[result["date"] == pd.Timestamp("2024-01-03")]
        assert len(jan3) == 1, "Jan 3 should be added"
        assert jan3["discharge"].values[0] == 30.0

    def test_should_reprocess_input_files_returns_true_when_no_output(self, tmp_path):
        """Test that should_reprocess_input_files returns True when output doesn't exist."""
        # Set up temp directories
        daily_dir = tmp_path / "daily_discharge"
        daily_dir.mkdir()
        intermediate_dir = tmp_path / "intermediate"
        intermediate_dir.mkdir()

        # Create an input file
        input_file = daily_dir / "test.xlsx"
        input_file.write_text("dummy")

        # Save original env vars
        orig_daily = os.environ.get("ieasyforecast_daily_discharge_path")
        orig_intermediate = os.environ.get("ieasyforecast_intermediate_data_path")
        orig_file = os.environ.get("ieasyforecast_daily_discharge_file")

        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(daily_dir)
            os.environ["ieasyforecast_intermediate_data_path"] = str(intermediate_dir)
            os.environ["ieasyforecast_daily_discharge_file"] = "output.csv"

            # Output file doesn't exist, should return True
            result = src.should_reprocess_input_files()
            assert result is True, "Should return True when output file doesn't exist"
        finally:
            # Restore original env vars
            if orig_daily is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig_daily
            if orig_intermediate is None:
                os.environ.pop("ieasyforecast_intermediate_data_path", None)
            else:
                os.environ["ieasyforecast_intermediate_data_path"] = orig_intermediate
            if orig_file is None:
                os.environ.pop("ieasyforecast_daily_discharge_file", None)
            else:
                os.environ["ieasyforecast_daily_discharge_file"] = orig_file


# ---------------------------------------------------------------------------
# Tests for uzhm wide-matrix Excel reader
# ---------------------------------------------------------------------------

import openpyxl  # noqa: E402  (appended after existing imports block)


def _build_uzhm_xlsx(path, station_header, year_data):
    """
    Create a uzhm-format xlsx file at *path*.

    Args:
        path: pathlib.Path — destination file path.
        station_header: str — e.g. "16022 Syrdariya-Chinaz", placed in E1.
        year_data: dict mapping year (int) -> list of 31 rows, where each row
            is a list of 12 discharge values (index 0=Jan, …, 11=Dec).
            Use None for invalid / non-existing dates.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for year, daily_values in year_data.items():
        ws = wb.create_sheet(title=str(year))

        # Row 1: station header in E1 (column 5)
        ws.cell(row=1, column=5, value=station_header)

        # Row 2: empty

        # Row 3: "Day" label
        ws.cell(row=3, column=1, value="Day")

        # Row 4: month numbers 1-12 in cols B-M (columns 2-13)
        for m in range(1, 13):
            ws.cell(row=4, column=m + 1, value=m)

        # Rows 5-35: day rows (days 1-31)
        for day_idx, discharges in enumerate(daily_values, start=1):
            row_num = 4 + day_idx  # row 5 = day 1
            ws.cell(row=row_num, column=1, value=day_idx)
            for month_idx, val in enumerate(discharges, start=1):
                ws.cell(row=row_num, column=month_idx + 1, value=val)

        # Row 37 (after day 31 at row 36): summary row that must be ignored
        ws.cell(row=37, column=1, value="Average")
        for m in range(1, 13):
            ws.cell(row=37, column=m + 1, value=500.0)

    wb.save(path)


def _make_year_data(base_value=800.0):
    """Return 31 rows x 12 months of discharge, with None for invalid dates."""
    import calendar

    # We'll use a simple pattern: base_value + day + month as the discharge
    rows = []
    for day in range(1, 32):
        discharges = []
        for month in range(1, 13):
            max_day = calendar.monthrange(2000, month)[1]
            if day > max_day:
                discharges.append(None)
            else:
                discharges.append(base_value + day + month)
        rows.append(discharges)
    return rows


def _make_uzhm_fixture(tmp_path):
    """
    Create two uzhm xlsx files in tmp_path:
      - 16022.xlsx  (station "16022 Syrdariya-Chinaz", years 2000 & 2001)
      - 16198.xlsx  (station "16198 AmuDarya-Termez", years 2000 & 2001)
    Returns tmp_path.
    """
    # Build year_data for 2000 (leap year) and 2001 (non-leap)
    year_data_2000 = _make_year_data(800.0)
    year_data_2001 = _make_year_data(700.0)

    # Override day=1 values for station 16022 to match test_happy_path assertion:
    # date(2000, 1, 1) -> discharge 832.0
    # Formula: 800 + day(1) + month(1) = 802 by default; override to 832
    year_data_2000[0][0] = 832.0  # day=1, January

    _build_uzhm_xlsx(
        tmp_path / "16022.xlsx",
        "16022 Syrdariya-Chinaz",
        {2000: year_data_2000, 2001: year_data_2001},
    )
    _build_uzhm_xlsx(
        tmp_path / "16198.xlsx",
        "16198 AmuDarya-Termez",
        {2000: _make_year_data(600.0), 2001: _make_year_data(500.0)},
    )
    return tmp_path


class TestUzhmWideMatrixReader:
    """Unit tests for the uzhm wide-matrix Excel reader functions."""

    def test_happy_path(self, tmp_path):
        fixture_dir = _make_uzhm_fixture(tmp_path)
        filename = fixture_dir / "16022.xlsx"

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(filename),
            code_list=["16022"],
            date_col="date",
            discharge_col="discharge",
            name_col="name",
            code_col="code",
        )

        assert isinstance(result, pd.DataFrame)
        assert not result.empty

        # Expected columns
        for col in ["date", "discharge", "code", "name"]:
            assert col in result.columns, f"Missing column: {col}"

        # Specific value: Jan 1 2000 should be 832.0
        # Use pd.Timestamp for comparison — the reader now returns datetime64[ns]
        # values (pd.Timestamp), so comparing with datetime.date would not match.
        jan1 = result[result["date"] == pd.Timestamp(2000, 1, 1)]
        assert not jan1.empty, "Expected a row for 2000-01-01"
        assert float(jan1["discharge"].iloc[0]) == 832.0

        # Code is int
        assert result["code"].dtype.kind in ("i", "u"), "code column should be integer type"
        assert int(result["code"].iloc[0]) == 16022

        # Name
        assert result["name"].iloc[0] == "Syrdariya-Chinaz"

        # Dates are coercible to datetime
        pd.to_datetime(result["date"])

    def test_invalid_dates_excluded(self, tmp_path):
        fixture_dir = _make_uzhm_fixture(tmp_path)
        filename = fixture_dir / "16022.xlsx"

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(filename),
            code_list=["16022"],
        )

        # Feb 30 and Feb 31 are not valid calendar dates and must not appear.
        # We check by filtering on year/month/day rather than constructing
        # Python date objects (which would raise ValueError).
        dates = result["date"]
        feb_day30_2000 = result[
            (dates.apply(lambda d: d.year) == 2000)
            & (dates.apply(lambda d: d.month) == 2)
            & (dates.apply(lambda d: d.day) == 30)
        ]
        assert feb_day30_2000.empty, "Feb 30 2000 should not appear in output"

        feb_day30_2001 = result[
            (dates.apply(lambda d: d.year) == 2001)
            & (dates.apply(lambda d: d.month) == 2)
            & (dates.apply(lambda d: d.day) == 30)
        ]
        assert feb_day30_2001.empty, "Feb 30 2001 should not appear in output"

        # Apr 31 must never appear (April has 30 days)
        apr_day31 = result[
            (dates.apply(lambda d: d.month) == 4) & (dates.apply(lambda d: d.day) == 31)
        ]
        assert apr_day31.empty, "Apr 31 should not appear in output"

        # February 2000 is a leap year (29 days)
        feb_2000 = result[
            (result["date"].apply(lambda d: d.year) == 2000)
            & (result["date"].apply(lambda d: d.month) == 2)
        ]
        assert len(feb_2000) == 29, f"Expected 29 Feb rows for leap year 2000, got {len(feb_2000)}"

        # February 2001 is not a leap year (28 days)
        feb_2001 = result[
            (result["date"].apply(lambda d: d.year) == 2001)
            & (result["date"].apply(lambda d: d.month) == 2)
        ]
        assert len(feb_2001) == 28, f"Expected 28 Feb rows for 2001, got {len(feb_2001)}"

    def test_none_discharge_excluded(self, tmp_path):
        """Rows with None discharge values must not appear in output."""

        # Build a year with explicit None for Jan 15
        year_data = _make_year_data(800.0)
        year_data[14][0] = None  # day=15, January -> None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2000"
        ws.cell(row=1, column=1, value="16999 Test-River")
        for m in range(1, 13):
            ws.cell(row=4, column=m + 1, value=m)
        for day_idx, discharges in enumerate(year_data, start=1):
            ws.cell(row=4 + day_idx, column=1, value=day_idx)
            for month_idx, val in enumerate(discharges, start=1):
                ws.cell(row=4 + day_idx, column=month_idx + 1, value=val)
        xlpath = tmp_path / "16999.xlsx"
        wb.save(xlpath)

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(xlpath),
            code_list=["16999"],
        )

        # Jan 15 should be absent
        # Use pd.Timestamp for comparison — the reader now returns datetime64[ns] values.
        jan15 = result[result["date"] == pd.Timestamp(2000, 1, 15)]
        assert jan15.empty, "date(2000,1,15) with None discharge should be excluded"

        # Jan 14 should be present (it has a valid value)
        jan14 = result[result["date"] == pd.Timestamp(2000, 1, 14)]
        assert not jan14.empty, "date(2000,1,14) should be present"

    def test_code_not_in_code_list_returns_empty(self, tmp_path):
        fixture_dir = _make_uzhm_fixture(tmp_path)
        filename = fixture_dir / "16022.xlsx"

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(filename),
            code_list=["99999"],
        )

        assert isinstance(result, pd.DataFrame)
        assert result.empty

    def test_non_year_sheet_skipped(self, tmp_path):
        """A sheet named 'Summary' must not crash the reader or add data."""
        # Build a file that has a "Summary" sheet alongside valid year sheets
        wb = openpyxl.Workbook()
        ws_year = wb.active
        ws_year.title = "2000"
        ws_year.cell(row=1, column=1, value="16100 RiverX")
        for m in range(1, 13):
            ws_year.cell(row=4, column=m + 1, value=m)
        for day in range(1, 29):  # safe days only
            ws_year.cell(row=4 + day, column=1, value=day)
            for m in range(1, 13):
                ws_year.cell(row=4 + day, column=m + 1, value=100.0 + day + m)

        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.cell(row=1, column=1, value="This should be ignored")
        ws_summary.cell(row=5, column=1, value=1)
        ws_summary.cell(row=5, column=2, value=9999.0)

        xlpath = tmp_path / "16100.xlsx"
        wb.save(xlpath)

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(xlpath),
            code_list=["16100"],
        )

        assert isinstance(result, pd.DataFrame)
        assert not result.empty

        # The value 9999.0 from the Summary sheet must not appear
        assert 9999.0 not in result["discharge"].values, (
            "Discharge 9999.0 from 'Summary' sheet should be excluded"
        )

    def test_summary_rows_ignored(self, tmp_path):
        """The 'Average' summary row (col A = 'Average') must be excluded."""
        fixture_dir = _make_uzhm_fixture(tmp_path)
        filename = fixture_dir / "16022.xlsx"

        result = src.read_runoff_data_from_uzhm_wide_xlsx(
            str(filename),
            code_list=["16022"],
        )

        # The fixture inserts summary rows with discharge=500.0 at row 37.
        # After _make_year_data the actual data values are all > 800 for 2000
        # (base 800 + day + month, minimum = 800+1+1=802) so 500.0 can only
        # come from a summary row.
        assert 500.0 not in result["discharge"].values, (
            "Summary row discharge 500.0 must not appear in output"
        )

    def test_directory_scan_pure_digits_goes_to_wide_matrix(self, tmp_path):
        """A pure-digit stem (19001.xlsx) is routed to the wide-matrix reader."""
        _build_uzhm_xlsx(
            tmp_path / "19001.xlsx",
            "19001 Test-River",
            {2000: _make_year_data(800.0)},
        )

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            # The wide-matrix file must contribute rows for code 19001
            codes_in_result = set(result["code"].unique())
            assert 19001 in codes_in_result, (
                "Expected code 19001 from wide-matrix file '19001.xlsx' in result"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_directory_scan_system_suffix_goes_to_single_river(self, tmp_path):
        """A SYSTEM-suffix stem (19002_Demo_UZB_SYSTEM.xlsx) is routed to the single-river reader."""
        _make_single_river_fixture(tmp_path, code="19002", river_name="Demo")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19002"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            # The single-river file must contribute rows for code 19002
            codes_in_result = set(result["code"].unique())
            assert 19002 in codes_in_result, "Expected code 19002 from single-river file in result"
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_organization_router_uzhm(self, tmp_path):
        """_read_runoff_data_by_organization('uzhm', ...) returns valid data."""
        fixture_dir = _make_uzhm_fixture(tmp_path)

        orig_path = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(fixture_dir)
            result = src._read_runoff_data_by_organization(
                "uzhm",
                date_col="date",
                discharge_col="discharge",
                name_col="name",
                code_col="code",
                code_list=["16022"],
            )

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty
            assert "date" in result.columns
            assert "discharge" in result.columns
        finally:
            if orig_path is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig_path

    def test_organization_router_uzhm_mixed_formats(self, tmp_path):
        """Org router combines wide-matrix and single-river results correctly.

        A directory with one wide-matrix file (19001.xlsx) and one single-river
        file (19999_Demo_River_UZB_SYSTEM.xlsx) must produce a DataFrame that
        contains rows for both station codes with correct column names and
        datetime dtypes.
        """

        # Wide-matrix file: pure-digit stem
        _build_uzhm_xlsx(
            tmp_path / "19001.xlsx",
            "19001 Test-River",
            {2000: _make_year_data(800.0)},
        )
        # Single-river file: SYSTEM suffix
        _make_single_river_fixture(tmp_path, code="19999", river_name="Demo_River")

        orig_path = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src._read_runoff_data_by_organization(
                "uzhm",
                date_col="date",
                discharge_col="discharge",
                name_col="name",
                code_col="code",
                code_list=["19001", "19999"],
            )

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            for col in ("date", "discharge", "code", "name"):
                assert col in result.columns, f"Expected column '{col}' in result"

            codes_in_result = set(result["code"].unique())
            assert 19001 in codes_in_result, (
                "Expected code 19001 from wide-matrix file '19001.xlsx' in result"
            )
            assert 19999 in codes_in_result, "Expected code 19999 from single-river file in result"

            assert result["date"].dtype.kind == "M", (
                f"Expected datetime64 dtype for 'date' column, got {result['date'].dtype}"
            )
        finally:
            if orig_path is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig_path

    def test_organization_router_uzhm_logs_summary(self, tmp_path, caplog):
        """The dispatcher emits a summary INFO log mentioning both format counts.

        After processing a directory with one wide-matrix file and one
        single-river file, read_all_runoff_data_from_uzhm_excel must emit at
        least one INFO record matching the pattern
        ``uzhm xlsx ingest: <N> wide-matrix + <M> single-river .* rows`` where
        N=1 and M=1.
        """
        import logging  # noqa: F811 — already imported later in module; safe here
        import re as _re

        # Wide-matrix file: pure-digit stem
        _build_uzhm_xlsx(
            tmp_path / "19001.xlsx",
            "19001 Test-River",
            {2000: _make_year_data(800.0)},
        )
        # Single-river file: SYSTEM suffix
        _make_single_river_fixture(tmp_path, code="19999", river_name="Demo_River")

        orig_path = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            with caplog.at_level(logging.INFO):
                src.read_all_runoff_data_from_uzhm_excel(
                    date_col="date",
                    discharge_col="discharge",
                    name_col="name",
                    code_col="code",
                    code_list=["19001", "19999"],
                )
        finally:
            if orig_path is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig_path

        pattern = _re.compile(r"uzhm xlsx ingest:.*wide-matrix.*single-river.*rows", _re.IGNORECASE)
        matching_records = [
            r for r in caplog.records if r.levelno == logging.INFO and pattern.search(r.message)
        ]
        assert matching_records, (
            "Expected at least one INFO record matching "
            f"'{pattern.pattern}', but none found. "
            f"Captured log records: {[r.message for r in caplog.records]}"
        )

        # The matched record must mention count 1 for wide-matrix and 1 for single-river.
        summary_msg = matching_records[0].message
        count_pattern = _re.compile(r"(\d+)\s+wide-matrix\s*\+\s*(\d+)\s+single-river")
        m = count_pattern.search(summary_msg)
        assert m is not None, (
            f"Could not extract wide-matrix/single-river counts from: '{summary_msg}'"
        )
        wide_count = int(m.group(1))
        sr_count = int(m.group(2))
        assert wide_count == 1, (
            f"Expected 1 wide-matrix file in log, got {wide_count}. Message: '{summary_msg}'"
        )
        assert sr_count == 1, (
            f"Expected 1 single-river file in log, got {sr_count}. Message: '{summary_msg}'"
        )


# ---------------------------------------------------------------------------
# Fixture helper for single-river xlsx files
# ---------------------------------------------------------------------------


def _build_single_river_xlsx(path, code, river_name, year_data):
    """
    Create a single-river-format xlsx file at *path*.

    Mirrors the format produced by
    ``apps/preprocessing_runoff/dev_code/convert_daily_runoff.py``
    ``write_single_river_excel()`` and consumed by
    ``read_runoff_data_from_single_river_xlsx()``.

    Args:
        path: pathlib.Path — destination file path.  The caller is responsible
            for using a filename that follows the convention
            ``{code}_{river_name}_UZB_SYSTEM.xlsx`` (exactly 5-digit code,
            exactly 16-char suffix ``_UZB_SYSTEM.xlsx``).
        code: str — 5-digit station code string (used only for documentation;
            the reader extracts the code from the filename, not the workbook).
        river_name: str — station name (again for documentation only; extracted
            from the filename by the reader).
        year_data: dict mapping year (int) -> list of (date_str, discharge)
            tuples.  date_str must be in ``dd.mm.YYYY`` format.  discharge may
            be a float, the string ``"-"``, or None (empty cell).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for year, rows in year_data.items():
        ws = wb.create_sheet(title=str(year))
        # Header row — the reader uses header=0 and overrides column names, so
        # actual values here don't matter, but we write canonical names for
        # readability.
        ws.append(["date", "discharge"])
        for date_str, discharge in rows:
            ws.append([date_str, discharge])

    wb.save(path)


def _make_single_river_fixture(tmp_path, code="19999", river_name="Demo_River"):
    """
    Create a single single-river xlsx file in tmp_path.

    Filename follows the ``{code}_{river_name}_UZB_SYSTEM.xlsx`` convention.
    Returns the full path to the created file.
    """
    filename = f"{code}_{river_name}_UZB_SYSTEM.xlsx"
    path = tmp_path / filename

    year_data = {
        2000: [
            ("01.01.2000", 100.5),
            ("02.01.2000", 101.5),
            ("03.01.2000", 102.5),
        ],
        2001: [
            ("01.01.2001", 200.5),
            ("02.01.2001", 201.5),
        ],
    }
    _build_single_river_xlsx(path, code, river_name, year_data)
    return path


# ---------------------------------------------------------------------------
# Tests for read_runoff_data_from_single_river_xlsx (documenting existing
# behavior — all tests in this class MUST pass against current src.py)
# ---------------------------------------------------------------------------


import logging  # noqa: E402  (appended after existing imports block)


class TestSingleRiverReader:
    """Unit tests for the single-river xlsx reader (no existing tests before P1)."""

    def test_happy_path(self, tmp_path):
        """Two-year file: returns rows with correct dtypes and known values."""
        filepath = _make_single_river_fixture(tmp_path, code="19999", river_name="Demo_River")

        result = src.read_runoff_data_from_single_river_xlsx(str(filepath), code_list=["19999"])

        assert isinstance(result, pd.DataFrame)
        assert not result.empty

        # Column presence (order-independent)
        assert {"date", "discharge", "code", "name"} <= set(result.columns)

        # Date dtype must be datetime64
        assert result["date"].dtype.kind == "M", (
            f"Expected datetime64 dtype for date column, got {result['date'].dtype}"
        )

        # Code dtype must be integer
        assert result["code"].dtype.kind in ("i", "u"), (
            f"Expected integer dtype for code column, got {result['code'].dtype}"
        )

        # Discharge dtype must be float
        assert result["discharge"].dtype.kind == "f", (
            f"Expected float dtype for discharge column, got {result['discharge'].dtype}"
        )

        # Specific-row assertion: Jan 2 2000 should have discharge 101.5
        jan2 = result[result["date"] == pd.Timestamp("2000-01-02")]
        assert not jan2.empty, "Expected a row for 2000-01-02"
        assert float(jan2["discharge"].iloc[0]) == 101.5

        # Rows from both years are present
        years = result["date"].dt.year.unique()
        assert 2000 in years
        assert 2001 in years

        # Total row count: 3 + 2 = 5
        assert len(result) == 5

    def test_code_list_filter_excludes(self, tmp_path, caplog):
        """File with code 19999 is skipped when code_list contains 19001 only."""
        filepath = _make_single_river_fixture(tmp_path, code="19999", river_name="Demo")

        with caplog.at_level(logging.DEBUG, logger="src"):
            result = src.read_runoff_data_from_single_river_xlsx(str(filepath), code_list=["19001"])

        assert isinstance(result, pd.DataFrame)
        assert result.empty

        # A debug log mentioning "not in code_list" must be emitted
        assert any(
            "not in code_list" in record.message.lower()
            or "not in code_list" in record.getMessage().lower()
            for record in caplog.records
        ), (
            "Expected a debug log message mentioning 'not in code_list' when "
            "code 19999 is excluded by code_list=['19001']"
        )

    def test_code_list_filter_includes(self, tmp_path):
        """File with code 19999 is included when code_list=['19999']."""
        filepath = _make_single_river_fixture(tmp_path, code="19999", river_name="Demo")

        result = src.read_runoff_data_from_single_river_xlsx(str(filepath), code_list=["19999"])

        assert isinstance(result, pd.DataFrame)
        assert not result.empty

    def test_missing_values_handled(self, tmp_path):
        """Rows with '-' or empty discharge cells become NaN in output."""
        path = tmp_path / "19876_MissingTest_UZB_SYSTEM.xlsx"
        year_data = {
            2000: [
                ("01.01.2000", 50.5),
                ("02.01.2000", "-"),  # explicit dash — must become NaN
                ("03.01.2000", None),  # empty cell — must become NaN
                ("04.01.2000", 55.5),
            ],
        }
        _build_single_river_xlsx(path, "19876", "MissingTest", year_data)

        result = src.read_runoff_data_from_single_river_xlsx(str(path), code_list=["19876"])

        assert isinstance(result, pd.DataFrame)
        assert len(result) == 4  # all rows returned, even with NaN

        discharges = result.sort_values("date")["discharge"].values
        # Jan 1: 50.5
        assert float(discharges[0]) == 50.5
        # Jan 2: NaN (from "-")
        assert pd.isna(discharges[1]), "Expected NaN for '-' discharge"
        # Jan 3: NaN (from empty cell)
        assert pd.isna(discharges[2]), "Expected NaN for empty discharge cell"
        # Jan 4: 55.5
        assert float(discharges[3]) == 55.5

    def test_file_not_found(self, tmp_path):
        """Nonexistent file path raises FileNotFoundError."""
        nonexistent = str(tmp_path / "99999_NoFile_UZB_SYSTEM.xlsx")

        with pytest.raises(FileNotFoundError):
            src.read_runoff_data_from_single_river_xlsx(nonexistent, code_list=["99999"])


# ---------------------------------------------------------------------------
# Tests for the new dispatch behavior of read_all_runoff_data_from_uzhm_excel
# (these MUST fail against current src.py — they describe P2's new behavior)
# ---------------------------------------------------------------------------


class TestUzhmExcelDispatch:
    """
    Dispatch tests for read_all_runoff_data_from_uzhm_excel.

    ALL tests in this class are expected to FAIL against current src.py.
    They pin the behavior that P2 will implement.  Assertion-level failures
    (not import/collection errors) confirm the gap.
    """

    def _make_wide_fixture(self, tmp_path, code="19001"):
        """Create a minimal wide-matrix xlsx in tmp_path and return path."""
        year_data = {2000: _make_year_data(800.0)}
        _build_uzhm_xlsx(
            tmp_path / f"{code}.xlsx",
            f"{code} Test-River",
            year_data,
        )
        return tmp_path / f"{code}.xlsx"

    def _make_single_river_in_dir(self, directory, code="19999", river_name="Demo_River"):
        """Create a single-river xlsx in directory and return path."""
        filename = f"{code}_{river_name}_UZB_SYSTEM.xlsx"
        path = directory / filename
        year_data = {
            2000: [
                ("01.01.2000", 300.5),
                ("02.01.2000", 301.5),
            ],
        }
        _build_single_river_xlsx(path, code, river_name, year_data)
        return path

    def test_uzhm_excel_dispatches_single_river_format(self, tmp_path):
        """Dir has both formats; result must contain rows for both codes."""
        self._make_wide_fixture(tmp_path, code="19001")
        self._make_single_river_in_dir(tmp_path, code="19999", river_name="Demo_River")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001", "19999"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            codes_in_result = set(result["code"].unique())
            assert 19001 in codes_in_result, "Expected code 19001 (wide-matrix) in result"
            assert 19999 in codes_in_result, (
                "Expected code 19999 (single-river) in result — "
                "single-river dispatch not yet implemented (P2)"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_single_river_only(self, tmp_path):
        """Dir has only a single-river file; result has code 19999."""
        self._make_single_river_in_dir(tmp_path, code="19999", river_name="Demo_River")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19999"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            codes_in_result = set(result["code"].unique())
            assert 19999 in codes_in_result, (
                "Expected code 19999 (single-river) in result — "
                "single-river dispatch not yet implemented (P2)"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_skips_unknown_filename_pattern(self, tmp_path):
        """Dir has a valid wide-matrix file and a garbage-named file; no exception raised."""
        self._make_wide_fixture(tmp_path, code="19001")

        # Create a garbage file with wide-matrix *content* but unrecognised name
        garbage_path = tmp_path / "garbage_file.xlsx"
        _build_uzhm_xlsx(
            garbage_path,
            "19002 Garbage-River",
            {2000: _make_year_data(400.0)},
        )

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            # Should not raise; data for 19001 returned; garbage data (400-range) absent
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001", "19002"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            # Data from garbage file uses base_value=400 → discharge > 400
            # Data from wide-matrix 19001 uses base_value=800 → discharge > 800
            # Garbage file's code 19002 must NOT appear in result
            codes_in_result = set(result["code"].unique())
            assert 19002 not in codes_in_result, (
                "Garbage file 'garbage_file.xlsx' must be skipped — unknown filename pattern"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_ignores_excel_temp_files(self, tmp_path):
        """Temp file ~$19001.xlsx must be skipped; no duplicate rows from it."""
        self._make_wide_fixture(tmp_path, code="19001")

        # Create a zero-byte temp file (as Excel would create it)
        temp_file = tmp_path / "~$19001.xlsx"
        temp_file.write_bytes(b"")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            # Count rows with code 19001; should come from exactly one file
            n_rows = len(result[result["code"] == 19001])
            # Wide-matrix for year 2000 has at most 366 rows; temp file would double it
            # We assert no duplication by checking the result is sensible (< 500 rows)
            assert n_rows < 500, (
                f"Too many rows ({n_rows}) for code 19001 — temp file may have been read"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_date_dtype_is_datetime64(self, tmp_path):
        """Combined result from both formats must have datetime64 date column."""
        self._make_wide_fixture(tmp_path, code="19001")
        self._make_single_river_in_dir(tmp_path, code="19999", river_name="Demo_River")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001", "19999"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            assert result["date"].dtype.kind == "M", (
                f"Expected datetime64[ns] for date column, got {result['date'].dtype} — "
                "wide-matrix reader returns datetime.date objects (P2 Part A) and "
                "single-river dispatch not yet implemented (P2 Part B)"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_column_set(self, tmp_path):
        """Combined result must have exactly columns {date, discharge, code, name}."""
        self._make_wide_fixture(tmp_path, code="19001")
        self._make_single_river_in_dir(tmp_path, code="19999", river_name="Demo_River")

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["19001", "19999"])

            assert result is not None
            assert isinstance(result, pd.DataFrame)
            assert not result.empty

            assert set(result.columns) == {"date", "discharge", "code", "name"}, (
                f"Expected column set exactly {{'date', 'discharge', 'code', 'name'}}, "
                f"got {set(result.columns)} — "
                "single-river dispatch not yet implemented (P2)"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_warns_on_duplicate_code(self, tmp_path, caplog):
        """When both 19001.xlsx and 19001_Demo_UZB_SYSTEM.xlsx exist, warn about duplicate."""
        # Wide-matrix file for 19001
        self._make_wide_fixture(tmp_path, code="19001")

        # Single-river file for the same code 19001
        sr_path = tmp_path / "19001_Demo_UZB_SYSTEM.xlsx"
        year_data = {2000: [("01.01.2000", 500.0)]}
        _build_single_river_xlsx(sr_path, "19001", "Demo", year_data)

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            with caplog.at_level(logging.WARNING):
                src.read_all_runoff_data_from_uzhm_excel(code_list=["19001"])

            warning_messages = [
                r.getMessage() for r in caplog.records if r.levelno >= logging.WARNING
            ]
            assert any("19001" in msg for msg in warning_messages), (
                "Expected a WARNING mentioning code 19001 for duplicate across "
                "wide-matrix and single-river formats — "
                "duplicate detection not yet implemented (P2)"
            )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig

    def test_uzhm_excel_routes_six_digit_stem_to_neither(self, tmp_path):
        """A 6-digit-prefix single-river file must NOT be routed to any reader."""
        # 6-digit prefix: reader would silently extract '12345' from '123456_...'[:5]
        six_digit_path = tmp_path / "123456_Demo_UZB_SYSTEM.xlsx"
        year_data = {2000: [("01.01.2000", 999.0)]}
        _build_single_river_xlsx(six_digit_path, "123456", "Demo", year_data)

        orig = os.environ.get("ieasyforecast_daily_discharge_path")
        try:
            os.environ["ieasyforecast_daily_discharge_path"] = str(tmp_path)
            # Pass both possible codes; neither should yield data from this file
            result = src.read_all_runoff_data_from_uzhm_excel(code_list=["12345", "123456"])

            # Either None or empty DataFrame — the 6-digit file must be skipped
            if result is not None and not result.empty:
                # Verify no rows came from the 6-digit file (discharge 999.0)
                assert 999.0 not in result["discharge"].values, (
                    "File '123456_Demo_UZB_SYSTEM.xlsx' (6-digit prefix) must be "
                    "skipped — the \\d{5} width constraint in P2 must be enforced"
                )
                # Neither code must appear in result
                codes_in_result = set(result["code"].unique())
                assert 12345 not in codes_in_result and 123456 not in codes_in_result, (
                    "6-digit-prefix file must not route any code into the result"
                )
        finally:
            if orig is None:
                os.environ.pop("ieasyforecast_daily_discharge_path", None)
            else:
                os.environ["ieasyforecast_daily_discharge_path"] = orig
