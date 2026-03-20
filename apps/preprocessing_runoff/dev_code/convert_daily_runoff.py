"""One-time conversion of Uzbekistan daily runoff data to SAPPHIRE format.

Converts two source Excel files for station 10001 (Zeravshan - Inflow to
Rovatkhodzha) into the single-river Excel format that the
preprocessing_runoff adapter ``read_runoff_data_from_single_river_xlsx``
expects.

Source files
------------
1. ``Inflow_Rovatkhodzha_daily_2010_2023.xlsx``
   Wide/pivot layout: column A = day-of-year dates (using 2023 as reference
   year, with row 61 = Feb 29 for leap years), columns B-O = discharge per
   year (2023 down to 2010).

2. ``Zeravshan - Inflow to Rovatkhodzha.xlsx``
   Calendar layout: one sheet per year (2024, 2025, 2026).  Row 2 headers =
   ``Day, January, February, ..., December`` (note typos "Yanuary", "Jube"
   in source -- we use column position instead of header text).  Rows 3-33 =
   day-of-month discharge values.

Output
------
A single Excel file named ``{code}_{name}_SYSTEM_ID.xlsx`` with one sheet
per year.  Each sheet has two columns (header row + data):

* Column A -- date as ``dd.mm.YYYY`` string
* Column B -- discharge (float, or empty for missing)

This matches the format consumed by
``read_runoff_data_from_single_river_xlsx`` in ``preprocessing_runoff``.

Usage
-----
::

    cd apps/preprocessing_runoff
    uv run python dev_code/convert_daily_runoff.py \\
        --input-dir "/path/to/data_forecast_tools/daily_runoff" \\
        --output-dir "/path/to/data_forecast_tools/daily_runoff"

    # Dry run (print summary without writing):
    uv run python dev_code/convert_daily_runoff.py \\
        --input-dir "/path/to/daily_runoff" --dry-run

After running, the output file coexists with the source files in the same
directory.  The preprocessing_runoff adapter picks it up automatically
because its filename starts with a 5-digit code.
"""

from __future__ import annotations

import argparse
import calendar
import logging
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# --- Constants ---------------------------------------------------------------

STATION_CODE = "10001"
STATION_NAME = "Zeravshan_Inflow_Rovatkhodzha"

FILE_PIVOT = "Inflow_Rovatkhodzha_daily_2010_2023.xlsx"
FILE_CALENDAR = "Zeravshan - Inflow to Rovatkhodzha.xlsx"


# --- Readers -----------------------------------------------------------------


def read_pivot_file(path: Path) -> pd.DataFrame:
    """Read the wide/pivot file and return long-format DataFrame.

    Returns:
        DataFrame with columns ``date`` (datetime) and ``discharge`` (float).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Row 0 is the header: Date, 2023, 2022, ..., 2010, [None padding]
    header = rows[0]
    years = [int(y) for y in header[1:] if y is not None]

    records: list[dict] = []
    for row in rows[1:]:
        ref_date = row[0]
        if ref_date is None:
            continue
        month = ref_date.month
        day = ref_date.day

        for i, year in enumerate(years):
            discharge = row[1 + i]
            if discharge is None:
                continue

            # Skip Feb 29 for non-leap years
            if month == 2 and day == 29 and not calendar.isleap(year):
                continue

            records.append(
                {
                    "date": date(year, month, day),
                    "discharge": float(discharge),
                }
            )

    df = pd.DataFrame(records)
    logger.info(
        "Pivot file: %d records, years %s-%s",
        len(df),
        min(years),
        max(years),
    )
    return df


def read_calendar_file(path: Path) -> pd.DataFrame:
    """Read the calendar-format file and return long-format DataFrame.

    Returns:
        DataFrame with columns ``date`` (datetime) and ``discharge`` (float).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(min_row=1, values_only=True))

        # Row 0: year in cell A1
        year_val = sheet_rows[0][0]
        if year_val is None:
            logger.warning("Sheet '%s': no year in A1, skipping", sheet_name)
            continue
        year = int(year_val)

        # Rows 2+ (0-indexed): day-of-month data.  Row 1 is the header.
        for row in sheet_rows[2:]:
            day_val = row[0]
            if day_val is None:
                continue
            day = int(day_val)

            for month in range(1, 13):
                discharge = row[month]  # column index = month (1-based)
                if discharge is None:
                    continue

                # Validate that this day exists in this month/year
                max_day = calendar.monthrange(year, month)[1]
                if day > max_day:
                    continue

                records.append(
                    {
                        "date": date(year, month, day),
                        "discharge": float(discharge),
                    }
                )

    wb.close()

    df = pd.DataFrame(records)
    if not df.empty:
        years = sorted(df["date"].apply(lambda d: d.year).unique())
        logger.info(
            "Calendar file: %d records, years %s",
            len(df),
            ", ".join(str(y) for y in years),
        )
    return df


# --- Writer ------------------------------------------------------------------


def write_single_river_excel(df: pd.DataFrame, output_path: Path) -> None:
    """Write long-format DataFrame to the single-river Excel format.

    One sheet per year, each with header ``Date | Discharge`` and dates
    formatted as ``dd.mm.YYYY``.
    """
    df = df.sort_values("date").drop_duplicates(subset="date", keep="last")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    years = sorted(df["date"].apply(lambda d: d.year).unique())
    for year in years:
        ws = wb.create_sheet(title=str(year))
        ws.append(["Date", "Discharge"])

        year_df = df[df["date"].apply(lambda d: d.year) == year].sort_values("date")
        for _, row in year_df.iterrows():
            ws.append([row["date"].strftime("%d.%m.%Y"), row["discharge"]])

    wb.save(output_path)
    logger.info("Wrote %d years to %s", len(years), output_path)


# --- Main --------------------------------------------------------------------


def make_output_filename(
    code: str = STATION_CODE,
    name: str = STATION_NAME,
) -> str:
    """Build filename matching the single-river naming convention.

    The adapter extracts:
    * code from ``filename[:5]``
    * name from ``filename[6:-16]``

    So the suffix after the name must be exactly 16 characters including
    the ``.xlsx`` extension: ``_XXXXXXXXXXX.xlsx`` (1 + 10 + 5 = 16).
    """
    # adapter does filename[6:-16] to extract name, so the last 16 chars
    # (including .xlsx) are discarded.  We need exactly 16 chars after name.
    # "_UZB_SYSTEM_.xlsx" = 17 chars → name would lose 1 char.
    # Use exactly 16: "_UZB_SYSTEM.xlsx"
    suffix = "_UZB_SYSTEM.xlsx"  # len = 16
    return f"{code}_{name}{suffix}"


def convert(input_dir: Path, output_dir: Path, dry_run: bool = False) -> Path:
    """Run the full conversion pipeline.

    Returns:
        Path to the output file (or would-be path if dry_run).
    """
    pivot_path = input_dir / FILE_PIVOT
    calendar_path = input_dir / FILE_CALENDAR

    if not pivot_path.exists():
        raise FileNotFoundError(f"Pivot file not found: {pivot_path}")
    if not calendar_path.exists():
        raise FileNotFoundError(f"Calendar file not found: {calendar_path}")

    df_pivot = read_pivot_file(pivot_path)
    df_calendar = read_calendar_file(calendar_path)

    df = pd.concat([df_pivot, df_calendar], ignore_index=True)
    df = df.sort_values("date").drop_duplicates(subset="date", keep="last")

    # Summary
    years = sorted(df["date"].apply(lambda d: d.year).unique())
    print(f"Combined: {len(df)} daily records, {len(years)} years ({years[0]}-{years[-1]})")
    for year in years:
        n = (df["date"].apply(lambda d: d.year) == year).sum()
        print(f"  {year}: {n} days")

    missing = df["discharge"].isna().sum()
    if missing:
        print(f"  Missing discharge values: {missing}")

    output_file = output_dir / make_output_filename()

    if dry_run:
        print(f"\n[DRY RUN] Would write to: {output_file}")
        return output_file

    write_single_river_excel(df, output_file)
    print(f"\nOutput: {output_file}")
    return output_file


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert daily runoff source files to SAPPHIRE format.",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        required=True,
        help="Directory containing the two source Excel files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory (default: same as input-dir).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print summary statistics without writing output.",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    output_dir = args.output_dir or args.input_dir
    convert(args.input_dir, output_dir, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
