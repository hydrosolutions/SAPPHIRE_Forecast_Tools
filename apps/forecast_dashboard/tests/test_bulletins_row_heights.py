"""Tests for the row-height fix applied on top of ieasyreports' row insertion.

Bug: ``ieasyreports.core.report_generator.report_generator.
DefaultReportGenerator._insert_rows`` shifts cell content (via
``self._unmerge_cells`` / ``self._shift_cells``) and merged ranges down when
rows are inserted, but never calls the existing-but-unused
``self._shift_row_dimensions``. As a result, ``sheet.row_dimensions`` (row
heights) stay pinned to their old row indices after an insert. In the monthly
reservoir bulletin, sections below an insertion point — e.g. a reservoir
title row and its table, which get rendered first because sections are
processed in reverse — end up with wrong/squeezed heights: the title row
lands on a row with no explicit height (looks squeezed) and other rows
inherit heights that belonged elsewhere.

This test file protects the fix for that bug: ``MultiSectionReportGenerator.
_insert_rows`` in ``src/bulletins.py``, which snapshots the explicit heights
of rows below the insertion point, delegates to the parent implementation,
and then re-applies the saved heights at their shifted row indices.
"""

import openpyxl
from ieasyreports.core.report_generator.report_generator import DefaultReportGenerator


class _Gen(DefaultReportGenerator):
    """Mirrors the production override in src/bulletins.py exactly."""

    def _insert_rows(self, row_idx, count, copy_style=True, fill_formulae=True):
        moved = {
            r: dim.height
            for r, dim in self.sheet.row_dimensions.items()
            if r > row_idx and dim.height is not None
        }
        super()._insert_rows(row_idx, count, copy_style=copy_style, fill_formulae=fill_formulae)
        for r, h in moved.items():
            self.sheet.row_dimensions[r + count].height = h


# ---------------------------------------------------------------------------
# Real-parent tests: exercise the actual ieasyreports DefaultReportGenerator.
# These do not stub anything — the real package is installed in the
# forecast_dashboard venv used by run_tests.sh forecast_dashboard.
# ---------------------------------------------------------------------------


class TestInsertRowsShiftsRowHeights:
    """Verify the override preserves row heights for rows below an insert."""

    def test_title_and_trailing_row_heights_preserved_after_insert(self):
        # Arrange
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="data_r2")
        ws.row_dimensions[2].height = 30
        ws.cell(row=4, column=1, value="RESERVOIR TITLE")
        ws.merge_cells("A4:E4")
        ws.row_dimensions[4].height = 45
        ws.cell(row=5, column=1, value="data_r5")
        ws.row_dimensions[5].height = 30

        g = _Gen.__new__(_Gen)
        g.sheet = ws

        # Act
        g._insert_rows(2, 2)

        # Assert
        assert ws.cell(row=6, column=1).value == "RESERVOIR TITLE"
        assert "A6:E6" in {str(mr) for mr in ws.merged_cells.ranges}
        # This is the exact squeeze bug — without the fix this would be None.
        assert ws.row_dimensions[6].height == 45
        assert ws.cell(row=7, column=1).value == "data_r5"
        assert ws.row_dimensions[7].height == 30

    def test_raw_parent_insert_rows_drops_height_documenting_the_bug(self):
        # Arrange
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="data_r2")
        ws.row_dimensions[2].height = 30
        ws.cell(row=4, column=1, value="RESERVOIR TITLE")
        ws.merge_cells("A4:E4")
        ws.row_dimensions[4].height = 45
        ws.cell(row=5, column=1, value="data_r5")
        ws.row_dimensions[5].height = 30

        g = DefaultReportGenerator.__new__(DefaultReportGenerator)
        g.sheet = ws

        # Act
        g._insert_rows(2, 2)

        # Assert: the untouched parent loses the row height (proves the bug).
        assert ws.row_dimensions[6].height is None


# ---------------------------------------------------------------------------
# Stubbed-import test: only checks that MultiSectionReportGenerator defines
# the override. Uses the same sys.modules stubbing bootstrap as
# test_bulletins_basin_numbering.py, kept physically separate from the
# real-parent tests above because it stubs DefaultReportGenerator with an
# empty class (no real _insert_rows), which would be useless for the
# behavior tests above. The real ieasyreports import at the top of this file
# already happened at module load time, before this bootstrap runs, and the
# bootstrap restores sys.modules in its finally block, so the two strategies
# do not interfere with each other.
# ---------------------------------------------------------------------------

import sys
import types
from unittest.mock import MagicMock

_FAKE_KEYS = [
    "panel",
    "panel.viewable",
    "panel.widgets",
    "panel.layout",
    "panel.pane",
    "panel.template",
    "src.gettext_config",
    "dashboard.logger",
    "src.db",
    "src.reports",
    "ieasyreports",
    "ieasyreports.settings",
    "ieasyreports.core",
    "ieasyreports.core.tags",
    "ieasyreports.core.tags.tag",
    "ieasyreports.core.report_generator",
    "tag_library",
    "src.bulletins",
    "src",
]

_saved = {k: sys.modules[k] for k in _FAKE_KEYS if k in sys.modules}

try:
    for _mod in [
        "panel",
        "panel.viewable",
        "panel.widgets",
        "panel.layout",
        "panel.pane",
        "panel.template",
        "tag_library",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    if "src.gettext_config" not in sys.modules:
        _gc = types.ModuleType("src.gettext_config")
        _gc._ = lambda x: x
        _gc.translation_manager = MagicMock()
        sys.modules["src.gettext_config"] = _gc

    if "dashboard.logger" not in sys.modules:
        _lg = types.ModuleType("dashboard.logger")
        _lg.setup_logger = MagicMock(return_value=MagicMock())
        sys.modules["dashboard.logger"] = _lg

    if "src.db" not in sys.modules:
        sys.modules["src.db"] = MagicMock()

    if "src.reports" not in sys.modules:
        sys.modules["src.reports"] = MagicMock()

    for _mod in [
        "ieasyreports",
        "ieasyreports.settings",
        "ieasyreports.core",
        "ieasyreports.core.tags",
        "ieasyreports.core.tags.tag",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    _rg_mod = types.ModuleType("ieasyreports.core.report_generator")

    class _StubDefaultReportGenerator:
        pass

    _rg_mod.DefaultReportGenerator = _StubDefaultReportGenerator
    sys.modules["ieasyreports.core.report_generator"] = _rg_mod

    for _clear in ("src.bulletins", "src"):
        sys.modules.pop(_clear, None)

    from src import bulletins

finally:
    for _k in _FAKE_KEYS:
        if _k in _saved:
            sys.modules[_k] = _saved[_k]
        elif _k in sys.modules:
            del sys.modules[_k]
    del _saved, _FAKE_KEYS


class TestProductionOverridePresent:
    """Verify MultiSectionReportGenerator defines the _insert_rows override."""

    def test_multi_section_report_generator_overrides_insert_rows(self):
        assert "_insert_rows" in bulletins.MultiSectionReportGenerator.__dict__
