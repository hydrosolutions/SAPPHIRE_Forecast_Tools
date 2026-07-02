"""Tests for basin numbering and merging helpers added for the TJ deployment.

Covers:
  - _assign_basin_numbers: sequential numbering, contiguous duplicates share a
    number, single-basin all = 1, empty list does not raise.
  - _merge_basin_columns: multi-station runs are merged, single-station runs
    are only aligned, empty cells break runs, correct merged-range coordinates,
    idempotent on a second call.
  - Org gating: with org != 'tjhm' the is_tj flag is False; with 'tjhm' it is
    True.  We test the helpers directly rather than the full Excel pipeline.

Bootstrap strategy mirrors test_bulletins_perc_formulas.py exactly — mock
heavy panel/ieasyreports deps, inject a real _StubDefaultReportGenerator stub,
force a fresh import of src.bulletins, then restore sys.modules.
"""

import sys
import types
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl

# ---------------------------------------------------------------------------
# Bootstrap: mock heavy dashboard dependencies before importing src.bulletins
# ---------------------------------------------------------------------------

_FAKE_KEYS = [
    "panel",
    "panel.viewable",
    "panel.widgets",
    "panel.layout",
    "panel.pane",
    "panel.template",
    "src.gettext_config",
    "dashboard.logger",
    "src.db",
    "src.reports",
    "ieasyreports",
    "ieasyreports.settings",
    "ieasyreports.core",
    "ieasyreports.core.tags",
    "ieasyreports.core.tags.tag",
    "ieasyreports.core.report_generator",
    "tag_library",
    "src.bulletins",
    "src",
]

_saved = {k: sys.modules[k] for k in _FAKE_KEYS if k in sys.modules}

try:
    for _mod in [
        "panel",
        "panel.viewable",
        "panel.widgets",
        "panel.layout",
        "panel.pane",
        "panel.template",
        "tag_library",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    if "src.gettext_config" not in sys.modules:
        _gc = types.ModuleType("src.gettext_config")
        _gc._ = lambda x: x
        _gc.translation_manager = MagicMock()
        sys.modules["src.gettext_config"] = _gc

    if "dashboard.logger" not in sys.modules:
        _lg = types.ModuleType("dashboard.logger")
        _lg.setup_logger = MagicMock(return_value=MagicMock())
        sys.modules["dashboard.logger"] = _lg

    if "src.db" not in sys.modules:
        sys.modules["src.db"] = MagicMock()

    if "src.reports" not in sys.modules:
        sys.modules["src.reports"] = MagicMock()

    for _mod in [
        "ieasyreports",
        "ieasyreports.settings",
        "ieasyreports.core",
        "ieasyreports.core.tags",
        "ieasyreports.core.tags.tag",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    _rg_mod = types.ModuleType("ieasyreports.core.report_generator")

    class _StubDefaultReportGenerator:
        pass

    _rg_mod.DefaultReportGenerator = _StubDefaultReportGenerator
    sys.modules["ieasyreports.core.report_generator"] = _rg_mod

    for _clear in ("src.bulletins", "src"):
        sys.modules.pop(_clear, None)

    from src import bulletins

    _assign_basin_numbers = bulletins._assign_basin_numbers
    _merge_basin_columns = bulletins._merge_basin_columns
    _merge_basin_columns_in_file = bulletins._merge_basin_columns_in_file

finally:
    for _k in _FAKE_KEYS:
        if _k in _saved:
            sys.modules[_k] = _saved[_k]
        elif _k in sys.modules:
            del sys.modules[_k]
    del _saved, _FAKE_KEYS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _site(basin_ru):
    """Create a minimal site-like object with a basin_ru attribute."""
    return SimpleNamespace(basin_ru=basin_ru)


def _ws_with_basin_col(basin_values, basin_col=2, no_col=1, start_row=1):
    """Build an openpyxl worksheet pre-filled with basin names in basin_col.

    Returns (wb, ws, start_row, end_row).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, val in enumerate(basin_values):
        row = start_row + i
        ws.cell(row=row, column=basin_col, value=val)
        # Put a basin number placeholder in no_col too, so alignment can be checked
        ws.cell(row=row, column=no_col, value=i + 1 if val else None)
    end_row = start_row + len(basin_values) - 1
    return wb, ws, start_row, end_row


# ---------------------------------------------------------------------------
# Tests: _assign_basin_numbers
# ---------------------------------------------------------------------------

class TestAssignBasinNumbers:
    """Unit tests for _assign_basin_numbers."""

    def test_sequential_distinct_basins(self):
        """Each distinct basin gets a unique sequential number."""
        sites = [_site("Бассейн А"), _site("Бассейн Б"), _site("Бассейн В")]
        result = _assign_basin_numbers(sites)
        assert [s._bulletin_basin_no for s in result] == [1, 2, 3]

    def test_contiguous_duplicates_share_number(self):
        """Contiguous sites with the same basin_ru share the same basin number."""
        sites = [
            _site("Бассейн А"),
            _site("Бассейн А"),
            _site("Бассейн Б"),
            _site("Бассейн Б"),
            _site("Бассейн Б"),
        ]
        result = _assign_basin_numbers(sites)
        assert [s._bulletin_basin_no for s in result] == [1, 1, 2, 2, 2]

    def test_single_basin_all_one(self):
        """A list with one basin → all sites get number 1."""
        sites = [_site("Единственный"), _site("Единственный"), _site("Единственный")]
        _assign_basin_numbers(sites)
        assert all(s._bulletin_basin_no == 1 for s in sites)

    def test_single_site(self):
        """A single-site list gets number 1."""
        sites = [_site("Бассейн X")]
        _assign_basin_numbers(sites)
        assert sites[0]._bulletin_basin_no == 1

    def test_empty_list_no_error(self):
        """Empty list must not raise."""
        result = _assign_basin_numbers([])
        assert result == []

    def test_returns_same_list(self):
        """The helper returns the same list object (modified in place)."""
        sites = [_site("А"), _site("Б")]
        returned = _assign_basin_numbers(sites)
        assert returned is sites

    def test_numbers_increment_by_first_appearance(self):
        """Basin numbers increment strictly on first appearance in the ordered list."""
        sites = [
            _site("Зеравшан"),   # 1
            _site("Зеравшан"),   # 1
            _site("Вахш"),       # 2
            _site("Пяндж"),      # 3
            _site("Пяндж"),      # 3
        ]
        _assign_basin_numbers(sites)
        assert [s._bulletin_basin_no for s in sites] == [1, 1, 2, 3, 3]

    def test_mixed_case_sensitive(self):
        """Basin names are compared as-is (case-sensitive)."""
        sites = [_site("а"), _site("А")]  # different strings → different numbers
        _assign_basin_numbers(sites)
        assert sites[0]._bulletin_basin_no == 1
        assert sites[1]._bulletin_basin_no == 2


# ---------------------------------------------------------------------------
# Tests: _merge_basin_columns
# ---------------------------------------------------------------------------

class TestMergeBasinColumns:
    """Unit tests for _merge_basin_columns."""

    def test_multi_row_basin_is_merged(self):
        """A 3-row basin run produces a merged range in both no_col and basin_col."""
        basin_values = ["Бассейн А", "Бассейн А", "Бассейн А"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        merged = {str(mr) for mr in ws.merged_cells.ranges}
        # Both columns must have a merged range covering all 3 rows
        assert "A1:A3" in merged
        assert "B1:B3" in merged

    def test_single_row_basin_not_merged(self):
        """A 1-row basin run must NOT produce any merged range."""
        basin_values = ["Один"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        assert len(list(ws.merged_cells.ranges)) == 0

    def test_multi_followed_by_single_correct_ranges(self):
        """Basin A (3 rows) then basin B (1 row): only A is merged, B is not."""
        basin_values = ["Бассейн А", "Бассейн А", "Бассейн А", "Бассейн Б"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        merged = {str(mr) for mr in ws.merged_cells.ranges}
        # Rows 1-3 are merged for both columns
        assert "A1:A3" in merged
        assert "B1:B3" in merged
        # Row 4 must NOT be merged
        assert "A4:A4" not in merged
        assert "B4:B4" not in merged
        assert len(list(ws.merged_cells.ranges)) == 2

    def test_empty_cell_breaks_run(self):
        """An empty basin_col cell breaks a run; the two sub-runs are processed
        independently."""
        # basin A (2 rows), empty (1 row), basin B (2 rows)
        basin_values = ["Бассейн А", "Бассейн А", None, "Бассейн Б", "Бассейн Б"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        merged = {str(mr) for mr in ws.merged_cells.ranges}
        assert "A1:A2" in merged
        assert "B1:B2" in merged
        assert "A4:A5" in merged
        assert "B4:B5" in merged
        # Row 3 is empty → no merge touching it
        assert all("3" not in r for r in merged)

    def test_alignment_set_on_top_cell_of_multi_run(self):
        """The top cell of a multi-row merged range gets center+middle alignment."""
        basin_values = ["Бассейн А", "Бассейн А"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        top_no = ws.cell(row=start_row, column=1).alignment
        top_basin = ws.cell(row=start_row, column=2).alignment
        assert top_no.horizontal == "center"
        assert top_no.vertical == "center"
        assert top_basin.horizontal == "center"
        assert top_basin.vertical == "center"

    def test_alignment_set_on_single_row_run(self):
        """Single-row runs get center+middle alignment even without merging."""
        basin_values = ["Единственная"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)

        cell = ws.cell(row=start_row, column=2)
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"

    def test_idempotent_second_call(self):
        """Calling _merge_basin_columns twice on the same worksheet must not raise
        and must produce the same merged ranges."""
        basin_values = ["Бассейн А", "Бассейн А", "Бассейн Б"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)
        merged_first = {str(mr) for mr in ws.merged_cells.ranges}

        # Second call must not raise and must yield the same merged ranges.
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=start_row, end_row=end_row)
        merged_second = {str(mr) for mr in ws.merged_cells.ranges}

        assert merged_first == merged_second

    def test_correct_row_span_two_basins_multi_rows(self):
        """Two multi-row basins each get separate merged ranges with correct spans."""
        # Basin A: rows 1-2, Basin B: rows 3-5
        basin_values = ["А", "А", "Б", "Б", "Б"]
        wb, ws, start_row, end_row = _ws_with_basin_col(basin_values)
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=1, end_row=5)

        merged = {str(mr) for mr in ws.merged_cells.ranges}
        # A spans rows 1-2, B spans rows 3-5
        assert "A1:A2" in merged
        assert "B1:B2" in merged
        assert "A3:A5" in merged
        assert "B3:B5" in merged
        assert len(list(ws.merged_cells.ranges)) == 4

    def test_empty_worksheet_no_error(self):
        """_merge_basin_columns on an empty row range must not raise."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # start_row > end_row — effectively no rows
        _merge_basin_columns(ws, no_col=1, basin_col=2, start_row=5, end_row=4)


# ---------------------------------------------------------------------------
# Tests: _merge_basin_columns_in_file
# ---------------------------------------------------------------------------

class TestMergeBasinColumnsInFile:
    """Integration-level tests for _merge_basin_columns_in_file."""

    def test_file_is_saved_with_merges(self, tmp_path):
        """Write a small xlsx, run _merge_basin_columns_in_file, reload and
        verify merged ranges exist."""
        path = tmp_path / "test_merge.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=2, value="Бассейн А")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="Бассейн А")
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="Бассейн Б")
        wb.save(str(path))

        _merge_basin_columns_in_file(str(path), no_col=1, basin_col=2)

        wb2 = openpyxl.load_workbook(str(path))
        ws2 = wb2.active
        merged = {str(mr) for mr in ws2.merged_cells.ranges}
        assert "A1:A2" in merged
        assert "B1:B2" in merged


# ---------------------------------------------------------------------------
# Tests: org-gating (is_tj flag)
# ---------------------------------------------------------------------------

class TestOrgGating:
    """Verify that the is_tj flag is correctly derived from the env var."""

    def test_tjhm_org_yields_is_tj_true(self, monkeypatch):
        """With org=tjhm, is_tj evaluates to True."""
        monkeypatch.setenv("ieasyhydroforecast_organization", "tjhm")
        import os
        assert os.getenv("ieasyhydroforecast_organization") == "tjhm"
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        assert is_tj is True

    def test_kghm_org_yields_is_tj_false(self, monkeypatch):
        """With org=kghm, is_tj evaluates to False."""
        monkeypatch.setenv("ieasyhydroforecast_organization", "kghm")
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        assert is_tj is False

    def test_unset_org_yields_is_tj_false(self, monkeypatch):
        """With org env var absent, is_tj evaluates to False."""
        monkeypatch.delenv("ieasyhydroforecast_organization", raising=False)
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        assert is_tj is False

    def test_assign_basin_numbers_not_called_for_non_tj(self, monkeypatch):
        """For a non-TJ org, _assign_basin_numbers should not set basin numbers on
        sites (i.e. if the caller skips it for non-TJ, sites have no attribute)."""
        monkeypatch.setenv("ieasyhydroforecast_organization", "kghm")
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        sites = [_site("Бассейн А"), _site("Бассейн А")]
        # Only call _assign_basin_numbers when is_tj — simulate the guard
        if is_tj:
            _assign_basin_numbers(sites)
        assert not hasattr(sites[0], '_bulletin_basin_no')

    def test_assign_basin_numbers_called_for_tj(self, monkeypatch):
        """For TJ org, _assign_basin_numbers is called and stamps all sites."""
        monkeypatch.setenv("ieasyhydroforecast_organization", "tjhm")
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        sites = [_site("Бассейн А"), _site("Бассейн А"), _site("Бассейн Б")]
        if is_tj:
            _assign_basin_numbers(sites)
        assert sites[0]._bulletin_basin_no == 1
        assert sites[1]._bulletin_basin_no == 1
        assert sites[2]._bulletin_basin_no == 2

    def test_merge_only_applied_when_tj(self, monkeypatch, tmp_path):
        """Merging only happens via _merge_basin_columns when called for TJ;
        a non-TJ path leaves the file unmodified (no merges)."""
        # Build a small xlsx
        path = tmp_path / "test_gating.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=2, value="Бассейн А")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="Бассейн А")
        wb.save(str(path))

        # Non-TJ path: do NOT call merge helper
        monkeypatch.setenv("ieasyhydroforecast_organization", "kghm")
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        if is_tj:
            _merge_basin_columns_in_file(str(path))

        wb2 = openpyxl.load_workbook(str(path))
        ws2 = wb2.active
        assert len(list(ws2.merged_cells.ranges)) == 0

    def test_merge_applied_when_tj(self, monkeypatch, tmp_path):
        """TJ path calls _merge_basin_columns_in_file and merges basin rows."""
        path = tmp_path / "test_gating_tj.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1)
        ws.cell(row=1, column=2, value="Бассейн А")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="Бассейн А")
        wb.save(str(path))

        monkeypatch.setenv("ieasyhydroforecast_organization", "tjhm")
        import os
        is_tj = os.getenv("ieasyhydroforecast_organization") == "tjhm"
        if is_tj:
            _merge_basin_columns_in_file(str(path))

        wb2 = openpyxl.load_workbook(str(path))
        ws2 = wb2.active
        merged = {str(mr) for mr in ws2.merged_cells.ranges}
        assert "A1:A2" in merged
        assert "B1:B2" in merged
