"""Tests for the module-level numerify helpers and NumerifyingReportGenerator
in bulletins.py.

These cover the fix for non-TJ pentad/decad Excel bulletins showing Excel's
"number stored as text" warning: non-TJ bulletins render through
DefaultReportGenerator.generate_report, which never converted the
tag-written comma-decimal strings (e.g. "12,3") into real numeric cells.
NumerifyingReportGenerator hooks save_report to run the same conversion
used by MultiSectionReportGenerator._numerify_value_cells, but across the
whole rendered sheet using the tag->column map from data_tags_info.

Bootstrap strategy: mirror test_bulletins_numeric_cells.py exactly — mock
heavy Panel/ieasyreports dependencies, inject a minimal real
_StubDefaultReportGenerator, force a fresh import of src.bulletins, then
restore sys.modules in a finally block so this module does not leak stubs
to other test modules.
"""

import os
import sys
import types
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl

# ---------------------------------------------------------------------------
# Bootstrap: mock heavy dashboard dependencies before importing src.bulletins
# ---------------------------------------------------------------------------

_FAKE_KEYS = [
    "panel",
    "panel.viewable",
    "panel.widgets",
    "panel.layout",
    "panel.pane",
    "panel.template",
    "src.gettext_config",
    "dashboard.logger",
    "src.db",
    "src.reports",
    "ieasyreports",
    "ieasyreports.settings",
    "ieasyreports.core",
    "ieasyreports.core.tags",
    "ieasyreports.core.tags.tag",
    "ieasyreports.core.report_generator",
    "tag_library",
    "src.bulletins",
    "src",
]

_saved = {k: sys.modules[k] for k in _FAKE_KEYS if k in sys.modules}

try:
    for _mod in [
        "panel",
        "panel.viewable",
        "panel.widgets",
        "panel.layout",
        "panel.pane",
        "panel.template",
        "tag_library",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    if "src.gettext_config" not in sys.modules:
        _gc = types.ModuleType("src.gettext_config")
        _gc._ = lambda x: x
        _gc.translation_manager = MagicMock()
        sys.modules["src.gettext_config"] = _gc

    if "dashboard.logger" not in sys.modules:
        _lg = types.ModuleType("dashboard.logger")
        _lg.setup_logger = MagicMock(return_value=MagicMock())
        sys.modules["dashboard.logger"] = _lg

    if "src.db" not in sys.modules:
        sys.modules["src.db"] = MagicMock()

    if "src.reports" not in sys.modules:
        sys.modules["src.reports"] = MagicMock()

    for _mod in [
        "ieasyreports",
        "ieasyreports.settings",
        "ieasyreports.core",
        "ieasyreports.core.tags",
        "ieasyreports.core.tags.tag",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    # Provide a minimal real DefaultReportGenerator stub so that
    # MultiSectionReportGenerator(DefaultReportGenerator) and
    # NumerifyingReportGenerator(DefaultReportGenerator) resolve to proper
    # Python types (object.__new__ requires a real type, not a MagicMock).
    _rg_mod = types.ModuleType("ieasyreports.core.report_generator")

    class _StubDefaultReportGenerator:
        pass

    _rg_mod.DefaultReportGenerator = _StubDefaultReportGenerator
    sys.modules["ieasyreports.core.report_generator"] = _rg_mod

    # Force a fresh import of src.bulletins so we always get the version
    # that inherits from our real stub, regardless of which test module ran
    # before.
    for _clear in ("src.bulletins", "src"):
        sys.modules.pop(_clear, None)

    from src import bulletins

    NumerifyingReportGenerator = bulletins.NumerifyingReportGenerator

finally:
    for _k in _FAKE_KEYS:
        if _k in _saved:
            sys.modules[_k] = _saved[_k]
        elif _k in sys.modules:
            del sys.modules[_k]
    del _saved, _FAKE_KEYS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_data_tags_info(ws, col_map):
    """Build a data_tags_info list from a dict of {tag_name: col_number}.

    Each entry is {"tag": SimpleNamespace(name=...), "cell": ws.cell(...)}.
    Row 1 is used only to anchor the cell object; _numerify_sheet_value_cells
    only reads ``cell.column`` from these entries.
    """
    data_tags_info = []
    for name, col in col_map.items():
        data_tags_info.append(
            {
                "tag": SimpleNamespace(name=name),
                "cell": ws.cell(row=1, column=col),
            }
        )
    return data_tags_info


# ---------------------------------------------------------------------------
# _numerify_cell unit tests
# ---------------------------------------------------------------------------


class TestNumerifyCell:
    """Unit tests for the module-level bulletins._numerify_cell helper."""

    def _fresh_cell(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = value
        return ws["A1"]

    def test_one_decimal_comma_string(self):
        cell = self._fresh_cell("12,3")
        bulletins._numerify_cell(cell)
        assert cell.value == 12.3
        assert cell.number_format == "0.0"

    def test_two_decimal_comma_string(self):
        cell = self._fresh_cell("1,23")
        bulletins._numerify_cell(cell)
        assert cell.value == 1.23
        assert cell.number_format == "0.00"

    def test_integer_string(self):
        cell = self._fresh_cell("123")
        bulletins._numerify_cell(cell)
        assert cell.value == 123
        assert isinstance(cell.value, int)
        assert cell.number_format == "0"

    def test_zero_string(self):
        cell = self._fresh_cell("0")
        bulletins._numerify_cell(cell)
        assert cell.value == 0
        assert isinstance(cell.value, int)
        assert cell.number_format == "0"

    def test_blank_string_untouched(self):
        cell = self._fresh_cell("")
        bulletins._numerify_cell(cell)
        assert cell.value in ("", None)

    def test_dash_spaces_untouched(self):
        cell = self._fresh_cell(" - ")
        bulletins._numerify_cell(cell)
        assert cell.value == " - "

    def test_em_dash_untouched(self):
        cell = self._fresh_cell("—")
        bulletins._numerify_cell(cell)
        assert cell.value == "—"

    def test_nan_string_untouched(self):
        cell = self._fresh_cell("nan")
        bulletins._numerify_cell(cell)
        assert cell.value == "nan"

    def test_already_numeric_cell_untouched(self):
        cell = self._fresh_cell(99.9)
        bulletins._numerify_cell(cell)
        assert cell.value == 99.9

    def test_none_cell_untouched(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws["A1"]
        assert cell.value is None
        bulletins._numerify_cell(cell)
        assert cell.value is None

    def test_negative_one_decimal(self):
        cell = self._fresh_cell("-1,5")
        bulletins._numerify_cell(cell)
        assert cell.value == -1.5
        assert cell.number_format == "0.0"


# ---------------------------------------------------------------------------
# _numerify_sheet_value_cells tests
# ---------------------------------------------------------------------------


class TestNumerifySheetValueCells:
    """Unit tests for bulletins._numerify_sheet_value_cells."""

    def test_converts_value_columns_across_all_rows_leaves_others(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Column layout: A=QEXP (value), B=DELTA (value),
        # C=PERC_NORM (value), D=MODEL (non-value), E=DASH (non-value)
        col_map = {
            "QEXP": 1,
            "DELTA": 2,
            "PERC_NORM": 3,
            "MODEL": 4,
            "DASH": 5,
        }
        data_tags_info = _make_data_tags_info(ws, col_map)

        rows_data = {
            2: ("12,3", "1,5", "45", "TFT", "—"),
            3: ("0", "-2,25", "100", "LR", "—"),
        }
        for row, (qexp, delta, perc, model, dash) in rows_data.items():
            ws.cell(row=row, column=1, value=qexp)
            ws.cell(row=row, column=2, value=delta)
            ws.cell(row=row, column=3, value=perc)
            ws.cell(row=row, column=4, value=model)
            ws.cell(row=row, column=5, value=dash)

        bulletins._numerify_sheet_value_cells(ws, data_tags_info)

        # Row 2 value columns numerified
        assert ws.cell(row=2, column=1).value == 12.3
        assert ws.cell(row=2, column=1).number_format == "0.0"
        assert ws.cell(row=2, column=2).value == 1.5
        assert ws.cell(row=2, column=2).number_format == "0.0"
        assert ws.cell(row=2, column=3).value == 45
        assert isinstance(ws.cell(row=2, column=3).value, int)
        assert ws.cell(row=2, column=3).number_format == "0"

        # Row 3 value columns numerified
        assert ws.cell(row=3, column=1).value == 0
        assert isinstance(ws.cell(row=3, column=1).value, int)
        assert ws.cell(row=3, column=2).value == -2.25
        assert ws.cell(row=3, column=2).number_format == "0.00"
        assert ws.cell(row=3, column=3).value == 100
        assert isinstance(ws.cell(row=3, column=3).value, int)

        # Non-value columns left untouched on every row
        assert ws.cell(row=2, column=4).value == "TFT"
        assert ws.cell(row=3, column=4).value == "LR"
        assert ws.cell(row=2, column=5).value == "—"
        assert ws.cell(row=3, column=5).value == "—"

    def test_no_value_columns_is_noop(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = {"MODEL": 1, "DASH": 2}
        data_tags_info = _make_data_tags_info(ws, col_map)
        ws.cell(row=2, column=1, value="TFT")
        ws.cell(row=2, column=2, value="—")

        bulletins._numerify_sheet_value_cells(ws, data_tags_info)

        assert ws.cell(row=2, column=1).value == "TFT"
        assert ws.cell(row=2, column=2).value == "—"

    def test_empty_data_tags_info_no_exception(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        bulletins._numerify_sheet_value_cells(ws, [])  # must not raise


# ---------------------------------------------------------------------------
# NumerifyingReportGenerator.save_report integration test
# ---------------------------------------------------------------------------


class TestNumerifyingReportGeneratorSaveReport:
    """Integration test proving save_report numerifies before persisting."""

    def test_save_report_numerifies_value_cells_before_saving(self, tmp_path, monkeypatch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="12,3")  # QEXP -> value column
        ws.cell(row=2, column=2, value="45")  # PERC_NORM -> value column
        ws.cell(row=2, column=3, value="TFT")  # MODEL -> non-value column

        data_tags_info = [
            {
                "tag": SimpleNamespace(name="QEXP"),
                "cell": ws.cell(row=1, column=1),
            },
            {
                "tag": SimpleNamespace(name="PERC_NORM"),
                "cell": ws.cell(row=1, column=2),
            },
            {
                "tag": SimpleNamespace(name="MODEL"),
                "cell": ws.cell(row=1, column=3),
            },
        ]

        instance = object.__new__(NumerifyingReportGenerator)
        instance.sheet = ws
        instance.data_tags_info = data_tags_info
        instance.template = wb
        instance.reports_directory_path = str(tmp_path)
        instance.template_filename = "template.xlsx"

        # The bootstrap's _StubDefaultReportGenerator has no save_report
        # method (unlike the real DefaultReportGenerator), so
        # super().save_report(...) would raise AttributeError. Give the
        # stub a minimal real save_report for the duration of this test.
        def _stub_save_report(self, name, output_path):
            if output_path is None:
                output_path = self.reports_directory_path
            os.makedirs(output_path, exist_ok=True)
            if name is None:
                name = f"{self.template_filename.split('.xlsx')[0]}.xlsx"
            self.template.save(os.path.join(output_path, name))

        monkeypatch.setattr(
            bulletins.DefaultReportGenerator,
            "save_report",
            _stub_save_report,
            raising=False,
        )

        instance.save_report("out.xlsx", str(tmp_path))

        saved_path = tmp_path / "out.xlsx"
        assert saved_path.exists()

        reopened = openpyxl.load_workbook(str(saved_path))
        rws = reopened.active

        qexp_cell = rws.cell(row=2, column=1)
        assert qexp_cell.data_type != "s"
        assert isinstance(qexp_cell.value, (int, float))
        assert qexp_cell.value == 12.3
        assert qexp_cell.number_format == "0.0"

        perc_cell = rws.cell(row=2, column=2)
        assert perc_cell.data_type != "s"
        assert isinstance(perc_cell.value, (int, float))
        assert perc_cell.value == 45

        model_cell = rws.cell(row=2, column=3)
        assert model_cell.value == "TFT"


# ---------------------------------------------------------------------------
# VALUE_TAG_NAMES whitelist assertions
# ---------------------------------------------------------------------------


class TestValueTagNamesWhitelist:
    """Sanity checks that VALUE_TAG_NAMES includes pentad/decad value tags
    and excludes non-value tags (MODEL, DASH)."""

    def test_qexp_is_a_value_tag(self):
        assert "QEXP" in bulletins.VALUE_TAG_NAMES

    def test_delta_is_a_value_tag(self):
        assert "DELTA" in bulletins.VALUE_TAG_NAMES

    def test_model_is_not_a_value_tag(self):
        assert "MODEL" not in bulletins.VALUE_TAG_NAMES

    def test_dash_is_not_a_value_tag(self):
        assert "DASH" not in bulletins.VALUE_TAG_NAMES
