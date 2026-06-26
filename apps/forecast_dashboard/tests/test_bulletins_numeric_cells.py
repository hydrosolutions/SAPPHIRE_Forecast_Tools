"""Tests for MultiSectionReportGenerator._numerify_value_cells (bulletins.py).

The method must convert comma-decimal text strings written by the tag system
into real numeric cells, preserving the displayed precision via a number format
('0', '0.0', '0.00').  Non-value columns, blank cells, dash strings, and
already-numeric cells must be left untouched.

Bootstrap strategy: mirror test_bulletins_perc_formulas.py exactly — mock heavy
Panel/ieasyreports dependencies, inject a minimal real _StubDefaultReportGenerator,
force a fresh import of src.bulletins, then restore sys.modules in a finally block
so this module does not leak stubs to other test modules.
"""

import sys
import types
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Bootstrap: mock heavy dashboard dependencies before importing src.bulletins
# ---------------------------------------------------------------------------

# src.bulletins and src are force-cleared so we always get a fresh import
# with our stub DefaultReportGenerator, regardless of import order with other
# test modules that bootstrap with a MagicMock DefaultReportGenerator.
_FAKE_KEYS = [
    "panel",
    "panel.viewable",
    "panel.widgets",
    "panel.layout",
    "panel.pane",
    "panel.template",
    "src.gettext_config",
    "dashboard.logger",
    "src.db",
    "src.reports",
    "ieasyreports",
    "ieasyreports.settings",
    "ieasyreports.core",
    "ieasyreports.core.tags",
    "ieasyreports.core.tags.tag",
    "ieasyreports.core.report_generator",
    "tag_library",
    "src.bulletins",
    "src",
]

_saved = {k: sys.modules[k] for k in _FAKE_KEYS if k in sys.modules}

try:
    for _mod in [
        "panel",
        "panel.viewable",
        "panel.widgets",
        "panel.layout",
        "panel.pane",
        "panel.template",
        "tag_library",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    if "src.gettext_config" not in sys.modules:
        _gc = types.ModuleType("src.gettext_config")
        _gc._ = lambda x: x
        _gc.translation_manager = MagicMock()
        sys.modules["src.gettext_config"] = _gc

    if "dashboard.logger" not in sys.modules:
        _lg = types.ModuleType("dashboard.logger")
        _lg.setup_logger = MagicMock(return_value=MagicMock())
        sys.modules["dashboard.logger"] = _lg

    if "src.db" not in sys.modules:
        sys.modules["src.db"] = MagicMock()

    if "src.reports" not in sys.modules:
        sys.modules["src.reports"] = MagicMock()

    for _mod in [
        "ieasyreports",
        "ieasyreports.settings",
        "ieasyreports.core",
        "ieasyreports.core.tags",
        "ieasyreports.core.tags.tag",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    # Provide a minimal real DefaultReportGenerator stub so that
    # MultiSectionReportGenerator(DefaultReportGenerator) resolves to a proper
    # Python type (object.__new__ requires a real type, not a MagicMock).
    _rg_mod = types.ModuleType("ieasyreports.core.report_generator")

    class _StubDefaultReportGenerator:
        pass

    _rg_mod.DefaultReportGenerator = _StubDefaultReportGenerator
    sys.modules["ieasyreports.core.report_generator"] = _rg_mod

    # Force a fresh import of src.bulletins so we always get the version that
    # inherits from our real stub, regardless of which test module ran before.
    for _clear in ("src.bulletins", "src"):
        sys.modules.pop(_clear, None)

    from src import bulletins

    MultiSectionReportGenerator = bulletins.MultiSectionReportGenerator

finally:
    for _k in _FAKE_KEYS:
        if _k in _saved:
            sys.modules[_k] = _saved[_k]
        elif _k in sys.modules:
            del sys.modules[_k]
    del _saved, _FAKE_KEYS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_section_cells(ws, col_map):
    """Build a section_cells list from a dict of {tag_name: col_number}.

    Each entry is {"tag": SimpleNamespace(name=...), "cell": ws.cell(row=1, col=N)}.
    Row 1 is used only to anchor the cell object; the actual data row is passed
    separately to _numerify_value_cells.
    """
    section_cells = []
    for name, col in col_map.items():
        section_cells.append({
            "tag": SimpleNamespace(name=name),
            "cell": ws.cell(row=1, column=col),
        })
    return section_cells


def _make_gen(ws):
    """Instantiate MultiSectionReportGenerator without calling __init__."""
    gen = object.__new__(MultiSectionReportGenerator)
    gen.sheet = ws
    return gen


# Standard column layout matching the template:
# Q_MIN=C=3, Q_MAX=E=5, V_MIN=F=6, V_MAX=H=8, NORM=I=9, VNORM=J=10
STANDARD_COL_MAP = {
    "Q_MIN": 3,    # C
    "Q_MAX": 5,    # E
    "V_MIN": 6,    # F
    "V_MAX": 8,    # H
    "NORM": 9,     # I
    "VNORM": 10,   # J
}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestNumerifyValueCells:
    """Unit tests for MultiSectionReportGenerator._numerify_value_cells."""

    def _fresh(self):
        """Return a fresh (workbook, active-sheet, generator) tuple."""
        wb = openpyxl.Workbook()
        ws = wb.active
        gen = _make_gen(ws)
        return wb, ws, gen

    # ── Core conversion cases ─────────────────────────────────────────────

    def test_one_decimal_float_and_format(self):
        """'12,3' -> float 12.3 with number_format '0.0'."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = "12,3"
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value == 12.3
        assert ws["C7"].number_format == "0.0"

    def test_two_decimal_float_and_format(self):
        """'1,23' -> float 1.23 with number_format '0.00'."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["E7"] = "1,23"
        gen._numerify_value_cells(7, section_cells)
        assert ws["E7"].value == 1.23
        assert ws["E7"].number_format == "0.00"

    def test_integer_norm_and_format(self):
        """'123' -> int 123 with number_format '0'."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["I7"] = "123"
        gen._numerify_value_cells(7, section_cells)
        assert ws["I7"].value == 123
        assert isinstance(ws["I7"].value, int)
        assert ws["I7"].number_format == "0"

    def test_large_integer_vnorm_and_format(self):
        """'12312' -> int 12312 with number_format '0'."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["J7"] = "12312"
        gen._numerify_value_cells(7, section_cells)
        assert ws["J7"].value == 12312
        assert isinstance(ws["J7"].value, int)
        assert ws["J7"].number_format == "0"

    def test_multiple_value_columns_at_once(self):
        """All six value columns are converted in a single call."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = "12,3"
        ws["E7"] = "1,23"
        ws["F7"] = "456"
        ws["H7"] = "7,8"
        ws["I7"] = "123"
        ws["J7"] = "12312"
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value == 12.3
        assert ws["C7"].number_format == "0.0"
        assert ws["E7"].value == 1.23
        assert ws["E7"].number_format == "0.00"
        assert ws["F7"].value == 456
        assert isinstance(ws["F7"].value, int)
        assert ws["F7"].number_format == "0"
        assert ws["H7"].value == 7.8
        assert ws["H7"].number_format == "0.0"
        assert ws["I7"].value == 123
        assert ws["I7"].number_format == "0"
        assert ws["J7"].value == 12312
        assert ws["J7"].number_format == "0"

    # ── Zero ─────────────────────────────────────────────────────────────

    def test_zero_string_becomes_int_zero(self):
        """'0' -> int 0, number_format '0'."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = "0"
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value == 0
        assert isinstance(ws["C7"].value, int)
        assert ws["C7"].number_format == "0"

    # ── Leave-untouched cases ─────────────────────────────────────────────

    def test_blank_string_left_untouched(self):
        """Empty string in a value column must not raise and the cell is unchanged."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = ""
        gen._numerify_value_cells(7, section_cells)
        # Still empty (or None as openpyxl may store it), no exception raised
        assert ws["C7"].value in ("", None)

    def test_non_value_column_left_untouched(self):
        """A tag not in the value set (e.g. RIVER_NAME) must not be converted."""
        _, ws, gen = self._fresh()
        # Add a non-value tag to the section_cells alongside the standard ones
        col_map = dict(STANDARD_COL_MAP)
        col_map["RIVER_NAME"] = 2  # column B
        section_cells = _make_section_cells(ws, col_map)
        ws["B7"] = "Нарын"
        gen._numerify_value_cells(7, section_cells)
        assert ws["B7"].value == "Нарын"

    def test_dash_string_in_value_column_left_untouched(self):
        """' - ' (unparseable) in a value column must be left as-is, no exception."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = " - "
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value == " - "

    def test_already_numeric_cell_skipped(self):
        """A cell already holding a number (not a str) must be left untouched."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = 99.9
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value == 99.9  # unchanged

    def test_none_cell_skipped(self):
        """A cell holding None must be left untouched without raising."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        # Default cell value is None; don't write anything
        gen._numerify_value_cells(7, section_cells)
        assert ws["C7"].value is None

    # ── Different row numbers ─────────────────────────────────────────────

    def test_different_row_number(self):
        """Row number is honoured correctly (row 14 example)."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C14"] = "56,7"
        gen._numerify_value_cells(14, section_cells)
        assert ws["C14"].value == 56.7
        assert ws["C14"].number_format == "0.0"
        # Row 7 is untouched
        assert ws["C7"].value is None

    def test_row_1_works(self):
        """Row 1 (minimum) is handled correctly."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C1"] = "8,90"
        gen._numerify_value_cells(1, section_cells)
        assert ws["C1"].value == 8.9
        assert ws["C1"].number_format == "0.00"

    # ── Empty section_cells ───────────────────────────────────────────────

    def test_empty_section_cells_no_exception(self):
        """Empty section_cells list must not raise."""
        _, ws, gen = self._fresh()
        gen._numerify_value_cells(7, [])  # must not raise

    # ── NaN guard ────────────────────────────────────────────────────────

    def test_nan_string_in_value_column_left_untouched(self):
        """'nan' in a value column must not raise and the cell is left as-is.

        Before the fix, float('nan') parsed successfully and then
        int(round(nan)) raised ValueError.  The guard skips NaN just like
        unparseable strings.
        """
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        ws["C7"] = "nan"
        gen._numerify_value_cells(7, section_cells)
        # Cell must be left unchanged — no conversion, no error
        assert ws["C7"].value == "nan"


# ---------------------------------------------------------------------------
# Tests for round_discharge_to_comma_separated_string
# ---------------------------------------------------------------------------

class TestRoundDischargeToCommaSeparatedString:
    """Unit tests for the round_discharge_to_comma_separated_string helper."""

    def test_nan_returns_empty_string(self):
        """float('nan') must return '' (blank cell), never the string 'nan'."""
        result = bulletins.round_discharge_to_comma_separated_string(float("nan"))
        assert result == ""

    def test_negative_returns_space(self):
        """Negative value returns a single space (existing behaviour unchanged)."""
        result = bulletins.round_discharge_to_comma_separated_string(-1.0)
        assert result == " "

    def test_zero_returns_zero_string(self):
        """Zero value returns '0' (existing behaviour unchanged)."""
        result = bulletins.round_discharge_to_comma_separated_string(0.0)
        assert result == "0"

    def test_small_value_two_decimals(self):
        """Value in (0, 10) formatted to 2 decimal places with comma separator."""
        result = bulletins.round_discharge_to_comma_separated_string(1.23)
        assert result == "1,23"

    def test_medium_value_one_decimal(self):
        """Value in [10, 100) formatted to 1 decimal place with comma separator."""
        result = bulletins.round_discharge_to_comma_separated_string(12.3)
        assert result == "12,3"

    def test_large_value_zero_decimals(self):
        """Value >= 100 formatted to 0 decimal places."""
        result = bulletins.round_discharge_to_comma_separated_string(100.1)
        assert result == "100"
