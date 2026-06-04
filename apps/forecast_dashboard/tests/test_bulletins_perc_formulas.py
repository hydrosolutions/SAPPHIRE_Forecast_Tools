"""Tests for MultiSectionReportGenerator._write_perc_formulas (bulletins.py).

The method must write live Excel IFERROR/ROUND formulas into the PERC_NORM (K)
and PERC_PREVYEAR (L) cells, deriving the column letters from the section's
own tag-cell mapping rather than hardcoding C/E/I/J/K/L.

Bootstrap strategy: mock heavy Panel/ieasyreports dependencies before importing
src.bulletins.  For ieasyreports.core.report_generator we supply a minimal real
stub class (DefaultReportGenerator) so that MultiSectionReportGenerator inherits
from a proper Python type and object.__new__() works in the tests.
"""

import sys
import types
from unittest.mock import MagicMock

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Bootstrap: mock heavy dashboard dependencies before importing src.bulletins
# ---------------------------------------------------------------------------

# Real openpyxl is already imported above; keep it in _saved so it survives
# the cleanup at the end of the bootstrap block.
# src.bulletins and src are force-cleared so we always get a fresh import
# with our stub DefaultReportGenerator, regardless of import order with other
# test modules that bootstrap with a MagicMock DefaultReportGenerator.
_FAKE_KEYS = [
    "panel",
    "panel.viewable",
    "panel.widgets",
    "panel.layout",
    "panel.pane",
    "panel.template",
    "src.gettext_config",
    "dashboard.logger",
    "src.db",
    "src.reports",
    "ieasyreports",
    "ieasyreports.settings",
    "ieasyreports.core",
    "ieasyreports.core.tags",
    "ieasyreports.core.tags.tag",
    "ieasyreports.core.report_generator",
    "tag_library",
    "src.bulletins",
    "src",
]

_saved = {k: sys.modules[k] for k in _FAKE_KEYS if k in sys.modules}

try:
    for _mod in [
        "panel",
        "panel.viewable",
        "panel.widgets",
        "panel.layout",
        "panel.pane",
        "panel.template",
        "tag_library",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    if "src.gettext_config" not in sys.modules:
        _gc = types.ModuleType("src.gettext_config")
        _gc._ = lambda x: x
        _gc.translation_manager = MagicMock()
        sys.modules["src.gettext_config"] = _gc

    if "dashboard.logger" not in sys.modules:
        _lg = types.ModuleType("dashboard.logger")
        _lg.setup_logger = MagicMock(return_value=MagicMock())
        sys.modules["dashboard.logger"] = _lg

    if "src.db" not in sys.modules:
        sys.modules["src.db"] = MagicMock()

    if "src.reports" not in sys.modules:
        sys.modules["src.reports"] = MagicMock()

    for _mod in [
        "ieasyreports",
        "ieasyreports.settings",
        "ieasyreports.core",
        "ieasyreports.core.tags",
        "ieasyreports.core.tags.tag",
    ]:
        if _mod not in sys.modules:
            sys.modules[_mod] = MagicMock()

    # Provide a minimal real DefaultReportGenerator stub so that
    # MultiSectionReportGenerator(DefaultReportGenerator) resolves to a proper
    # Python type (object.__new__ requires a real type, not a MagicMock).
    _rg_mod = types.ModuleType("ieasyreports.core.report_generator")

    class _StubDefaultReportGenerator:
        pass

    _rg_mod.DefaultReportGenerator = _StubDefaultReportGenerator
    sys.modules["ieasyreports.core.report_generator"] = _rg_mod

    # Force a fresh import of src.bulletins so we always get the version that
    # inherits from our real stub, regardless of which test module ran before.
    for _clear in ("src.bulletins", "src"):
        sys.modules.pop(_clear, None)

    from src import bulletins

    MultiSectionReportGenerator = bulletins.MultiSectionReportGenerator

finally:
    for _k in _FAKE_KEYS:
        if _k in _saved:
            sys.modules[_k] = _saved[_k]
        elif _k in sys.modules:
            del sys.modules[_k]
    del _saved, _FAKE_KEYS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_section_cells(ws, col_map):
    """Build a section_cells list from a dict of {tag_name: col_number}.

    Each entry is {"tag": SimpleNamespace(name=...), "cell": ws.cell(row=1, col=N)}.
    Row 1 is used only to anchor the cell object; the actual data row is passed
    separately to _write_perc_formulas.
    """
    section_cells = []
    for name, col in col_map.items():
        section_cells.append({
            "tag": types.SimpleNamespace(name=name),
            "cell": ws.cell(row=1, column=col),
        })
    return section_cells


def _make_gen(ws):
    """Instantiate MultiSectionReportGenerator without calling __init__."""
    gen = object.__new__(MultiSectionReportGenerator)
    gen.sheet = ws
    return gen


# Standard column layout that matches the template (Q_MIN=C=3, Q_MAX=E=5,
# NORM=I=9, VNORM=J=10, PERC_NORM=K=11, PERC_PREVYEAR=L=12).
STANDARD_COL_MAP = {
    "Q_MIN": 3,    # C
    "Q_MAX": 5,    # E
    "NORM": 9,     # I
    "VNORM": 10,   # J
    "PERC_NORM": 11,   # K
    "PERC_PREVYEAR": 12,  # L
}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestWritePercFormulas:
    """Unit tests for MultiSectionReportGenerator._write_perc_formulas."""

    def _fresh(self):
        """Return a fresh (workbook, active-sheet, generator) tuple."""
        wb = openpyxl.Workbook()
        ws = wb.active
        gen = _make_gen(ws)
        return wb, ws, gen

    # ── Basic formula content ─────────────────────────────────────────────

    def test_k7_formula_standard_columns(self):
        """K7 gets the correct IFERROR/ROUND formula for the standard layout."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        gen._write_perc_formulas(7, section_cells)
        assert ws["K7"].value == '=IFERROR(ROUND((C7+E7)/2/I7*100,0),"")'

    def test_l7_formula_standard_columns(self):
        """L7 gets the correct IFERROR/ROUND formula for the standard layout."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        gen._write_perc_formulas(7, section_cells)
        assert ws["L7"].value == '=IFERROR(ROUND((C7+E7)/2/J7*100,0),"")'

    # ── Different row numbers ─────────────────────────────────────────────

    def test_row_14_produces_row_14_references(self):
        """Formulas for row 14 must reference row 14, not row 7."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        gen._write_perc_formulas(14, section_cells)
        assert ws["K14"].value == '=IFERROR(ROUND((C14+E14)/2/I14*100,0),"")'
        assert ws["L14"].value == '=IFERROR(ROUND((C14+E14)/2/J14*100,0),"")'

    def test_row_1_produces_row_1_references(self):
        """Boundary: first data row (1) uses row number 1 in references."""
        _, ws, gen = self._fresh()
        section_cells = _make_section_cells(ws, STANDARD_COL_MAP)
        gen._write_perc_formulas(1, section_cells)
        assert ws["K1"].value == '=IFERROR(ROUND((C1+E1)/2/I1*100,0),"")'
        assert ws["L1"].value == '=IFERROR(ROUND((C1+E1)/2/J1*100,0),"")'

    # ── Missing tags — partial write + no exception ───────────────────────

    def test_norm_absent_k_untouched_no_exception(self):
        """When NORM is absent, K stays None and no exception is raised."""
        _, ws, gen = self._fresh()
        col_map_no_norm = {k: v for k, v in STANDARD_COL_MAP.items() if k != "NORM"}
        section_cells = _make_section_cells(ws, col_map_no_norm)
        gen._write_perc_formulas(7, section_cells)
        assert ws["K7"].value is None  # untouched
        # L should still be written (VNORM is present)
        assert ws["L7"].value == '=IFERROR(ROUND((C7+E7)/2/J7*100,0),"")'

    def test_perc_prevyear_absent_l_untouched_no_exception(self):
        """When PERC_PREVYEAR tag is absent, L is untouched and no exception is raised."""
        _, ws, gen = self._fresh()
        col_map_no_l = {k: v for k, v in STANDARD_COL_MAP.items() if k != "PERC_PREVYEAR"}
        section_cells = _make_section_cells(ws, col_map_no_l)
        gen._write_perc_formulas(7, section_cells)
        # K must still be written
        assert ws["K7"].value == '=IFERROR(ROUND((C7+E7)/2/I7*100,0),"")'
        # L column cell should be None (never written)
        assert ws["L7"].value is None

    def test_vnorm_absent_l_untouched_no_exception(self):
        """When VNORM is absent, L stays None."""
        _, ws, gen = self._fresh()
        col_map_no_vnorm = {k: v for k, v in STANDARD_COL_MAP.items() if k != "VNORM"}
        section_cells = _make_section_cells(ws, col_map_no_vnorm)
        gen._write_perc_formulas(7, section_cells)
        assert ws["L7"].value is None

    def test_all_perc_tags_absent_no_exception(self):
        """Completely empty section_cells raises no exception and writes nothing."""
        _, ws, gen = self._fresh()
        gen._write_perc_formulas(7, [])
        assert ws["K7"].value is None
        assert ws["L7"].value is None

    # ── Column letters derived from actual tag columns ─────────────────────

    def test_column_letters_derived_from_actual_columns(self):
        """Shift NORM to a different column; formula must use that column's letter."""
        _, ws, gen = self._fresh()
        # Move NORM from col 9 (I) to col 15 (O)
        shifted_col_map = dict(STANDARD_COL_MAP)
        shifted_col_map["NORM"] = 15  # O
        section_cells = _make_section_cells(ws, shifted_col_map)
        gen._write_perc_formulas(7, section_cells)
        # K7 must reference O, not I
        assert ws["K7"].value == '=IFERROR(ROUND((C7+E7)/2/O7*100,0),"")'

    def test_column_letters_derived_qmin_qmax_shift(self):
        """Shift Q_MIN/Q_MAX; both K and L formulas must use the new letters."""
        _, ws, gen = self._fresh()
        shifted_col_map = dict(STANDARD_COL_MAP)
        shifted_col_map["Q_MIN"] = 2   # B
        shifted_col_map["Q_MAX"] = 4   # D
        section_cells = _make_section_cells(ws, shifted_col_map)
        gen._write_perc_formulas(7, section_cells)
        assert ws["K7"].value == '=IFERROR(ROUND((B7+D7)/2/I7*100,0),"")'
        assert ws["L7"].value == '=IFERROR(ROUND((B7+D7)/2/J7*100,0),"")'

    def test_perc_norm_column_derived_from_tag(self):
        """If PERC_NORM is mapped to a column other than K, formula goes there."""
        _, ws, gen = self._fresh()
        shifted_col_map = dict(STANDARD_COL_MAP)
        shifted_col_map["PERC_NORM"] = 13  # M
        section_cells = _make_section_cells(ws, shifted_col_map)
        gen._write_perc_formulas(7, section_cells)
        # Formula must land in M7, not K7
        assert ws["M7"].value == '=IFERROR(ROUND((C7+E7)/2/I7*100,0),"")'
        assert ws["K7"].value is None  # original K untouched
