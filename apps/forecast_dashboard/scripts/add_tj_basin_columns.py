"""
add_tj_basin_columns.py
-----------------------
Idempotent openpyxl transform that inserts two leading columns (№, Бассейн)
into the four Tajik bulletin Excel template files for SAPPHIRE Forecast Tools.

Design reference:
  doc/plans/tj_bulletin_basin_numbering_design.md

Target files (Tajik copies ONLY):
  taj_data_forecast_tools/templates/pentadal_forecast_bulletin_template_tj.xlsx
  taj_data_forecast_tools/templates/decadal_forecast_bulletin_template_tj.xlsx
  taj_data_forecast_tools/templates/monthly_forecast_bulletin_template_tj.xlsx
  taj_data_forecast_tools/templates/seasonal_forecast_bulletin_template_tj.xlsx

Idempotency guard: if column A of the relevant header row already contains
'№', the file is skipped.

Backup policy: each file is backed up to <name>.bak before editing; if a .bak
already exists the backup step is skipped (to protect the original on reruns).

Key openpyxl behaviour note:
  ws.insert_cols() shifts cell values but does NOT update merged_cells.ranges.
  The correct approach is:
    1. Snapshot all merge ranges as plain (min_col, max_col, min_row, max_row)
       tuples.
    2. Unmerge ALL ranges before inserting.
    3. Call insert_cols — now cells shift correctly with no stale merges.
    4. Re-create every merge with min_col/max_col incremented by N_INSERT.
    5. Add new A/B merges (header span rows 5-6 for pentadal, none for monthly).
    6. Extend title/separator merges to the new full width.

Usage:
    python apps/forecast_dashboard/scripts/add_tj_basin_columns.py

    # Or with an explicit templates directory:
    python apps/forecast_dashboard/scripts/add_tj_basin_columns.py \\
        --templates-dir /path/to/taj_data_forecast_tools/templates
"""

import argparse
import copy
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

# ---------------------------------------------------------------------------
# Default template directory
# ---------------------------------------------------------------------------
_DEFAULT_TMPL_DIR = (
    Path.home()
    / "hydrosolutions Dropbox"
    / "Maxat Pernebayev"
    / "SAPPHIRE_Central_Asia_Technical_Work"
    / "data"
    / "taj_data_forecast_tools"
    / "templates"
)

# Files to transform (Tajik copies only)
TJ_TEMPLATES = [
    "pentadal_forecast_bulletin_template_tj.xlsx",
    "decadal_forecast_bulletin_template_tj.xlsx",
    "monthly_forecast_bulletin_template_tj.xlsx",
    "seasonal_forecast_bulletin_template_tj.xlsx",
]

# How many columns to insert at the left
N_INSERT = 2


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------


def _copy_cell_style(src_cell, dst_cell) -> None:
    """Copy font, fill, border, and alignment from src to dst."""
    if src_cell.font:
        dst_cell.font = copy.copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy.copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy.copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format


def _snapshot_merges(ws) -> list[tuple[int, int, int, int]]:
    """
    Snapshot all merged ranges as plain (min_col, max_col, min_row, max_row)
    tuples. Must be called BEFORE insert_cols.
    """
    return [
        (mr.min_col, mr.max_col, mr.min_row, mr.max_row)
        for mr in ws.merged_cells.ranges
    ]


def _unmerge_all(ws) -> None:
    """Remove every merge from the worksheet."""
    for ms in [str(mr) for mr in ws.merged_cells.ranges]:
        ws.unmerge_cells(ms)


def _rebuild_merges(ws, snapshots: list[tuple], delta: int = N_INSERT) -> None:
    """
    Re-create every merge from snapshots with each column shifted by delta.

    All merges must have been removed with _unmerge_all() before insert_cols,
    and this function is called after insert_cols.
    """
    for min_col, max_col, min_row, max_row in snapshots:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col + delta,
            end_row=max_row,
            end_column=max_col + delta,
        )


def _set_title_merge_width(ws, row: int, new_max_col: int) -> None:
    """
    Find the full-width merge in `row` (which after _rebuild_merges starts at
    col 1+N_INSERT because all merges were shifted), extend it to A:new_max_col.

    After insert_cols, the original cell value moved to col 1+N_INSERT.  We
    move that value to A (col 1) before re-merging so the title is not lost.
    """
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == row and mr.max_row == row:
            shifted_origin_col = mr.min_col  # e.g. col 3 after +2 shift
            # Move value from shifted origin to col 1 (A) if it is not there
            if shifted_origin_col != 1:
                src = ws.cell(row, shifted_origin_col)
                dst = ws.cell(row, 1)
                dst.value = src.value
                src.value = None
                # Copy style too
                if src.font:
                    dst.font = copy.copy(src.font)
                if src.fill:
                    dst.fill = copy.copy(src.fill)
                if src.border:
                    dst.border = copy.copy(src.border)
                if src.alignment:
                    dst.alignment = copy.copy(src.alignment)
                if src.number_format:
                    dst.number_format = src.number_format
            ws.unmerge_cells(str(mr))
            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=new_max_col,
            )
            return


# ---------------------------------------------------------------------------
# Pentadal / decadal transform
# ---------------------------------------------------------------------------


def _transform_pentadal(ws) -> None:
    """
    Transform the '1 пентада' sheet.

    Before (cols A-S = 19, last used col 19):
      Row 1:  title  merged A1:S1
      Row 2:  subtitle merged A2:S2
      Row 4:  "Средний расход воды (факт)" A4:G4 | H=gap | "Прогноз водности" I4:S4
      Row 5:  headers (A=РЕКА, B=ПУНКТ, C-G=actual, I=РЕКА, J=ПУНКТ, K-S=forecast)
      Row 6:  sub-headers (Q6=мин, R6=норма, S6=макс)
      Row 7:  {{HEADER.BASIN_RU}} merged A7:S7
      Row 8:  {{DATA.*}} placeholders

    After (cols A-U = 21):
      Row 1:  title  merged A1:U1
      Row 2:  subtitle merged A2:U2
      Row 4:  section headers shifted to C4:I4 / K4:U4
      Row 5:  A5:A6=№, B5:B6=Бассейн (merged), C5=old A5 onward
      Row 7:  separator merged A7:U7, NO {{HEADER.*}} content
      Row 8:  A8={{DATA.BASIN_NO}}, B8={{DATA.BASIN_RU}}, C8=old A8 onward
    """
    # Old dimensions: last data column was S (col 19), new will be U (col 21)
    new_max_col = 21  # U

    # 1. Capture style reference cells BEFORE modifying anything
    # old A5 = "РЕКА" header — used as style source for №/Бассейн headers
    ref_hdr = ws["A5"]
    ref_hdr_font = copy.copy(ref_hdr.font)
    ref_hdr_fill = copy.copy(ref_hdr.fill)
    ref_hdr_border = copy.copy(ref_hdr.border)
    ref_hdr_number_format = ref_hdr.number_format

    # old A8 = "{{DATA.RIVER_NAME_RU}}" — used as style source for data cells
    ref_data = ws["A8"]
    ref_data_font = copy.copy(ref_data.font)
    ref_data_fill = copy.copy(ref_data.fill)
    ref_data_border = copy.copy(ref_data.border)
    ref_data_number_format = ref_data.number_format

    # 2. Snapshot all merges as plain tuples
    snapshots = _snapshot_merges(ws)

    # 3. Unmerge ALL before insert (openpyxl does not shift merges on insert)
    _unmerge_all(ws)

    # 4. Insert 2 columns at the far left
    ws.insert_cols(1, N_INSERT)
    # Now: old col N -> new col N+2; A-B are empty new columns

    # 5. Rebuild all original merges shifted by +2
    _rebuild_merges(ws, snapshots, delta=N_INSERT)
    # Now merges are e.g. C1:U1 (was A1:S1), C7:U7 (was A7:S7), etc.

    # 6. Extend title rows 1 and 2 to the full new width (A1:U1, A2:U2)
    # These are currently at C1:U1 / C2:U2 after shift; extend to A1.
    _set_title_merge_width(ws, 1, new_max_col)
    _set_title_merge_width(ws, 2, new_max_col)

    # 7. Fix the row-7 separator: was A7:S7 → now C7:U7; extend to A7:U7
    #    Also clear the {{HEADER.BASIN_RU}} content (now in C7 after shift)
    ws["C7"].value = None
    ws["A7"].value = None
    _set_title_merge_width(ws, 7, new_max_col)

    # 8. Add new header cells A5:A6 = № and B5:B6 = Бассейн
    #    First remove any stale merge that may occupy A5 or B5
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= 5 <= mr.max_row and mr.min_col in (1, 2):
            ws.unmerge_cells(str(mr))

    ws["A5"].value = "№"
    ws["A5"].font = ref_hdr_font
    ws["A5"].fill = ref_hdr_fill
    ws["A5"].border = ref_hdr_border
    ws["A5"].number_format = ref_hdr_number_format
    ws["A5"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)

    ws["B5"].value = "Бассейн"
    ws["B5"].font = ref_hdr_font
    ws["B5"].fill = ref_hdr_fill
    ws["B5"].border = ref_hdr_border
    ws["B5"].number_format = ref_hdr_number_format
    ws["B5"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)

    # 9. Add new data cells A8 = {{DATA.BASIN_NO}}, B8 = {{DATA.BASIN_RU}}
    ws["A8"].value = "{{DATA.BASIN_NO}}"
    ws["A8"].font = ref_data_font
    ws["A8"].fill = ref_data_fill
    ws["A8"].border = ref_data_border
    ws["A8"].number_format = ref_data_number_format
    ws["A8"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws["B8"].value = "{{DATA.BASIN_RU}}"
    ws["B8"].font = ref_data_font
    ws["B8"].fill = ref_data_fill
    ws["B8"].border = ref_data_border
    ws["B8"].number_format = ref_data_number_format
    ws["B8"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 10. Set column widths for the two new columns
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 18.0


# ---------------------------------------------------------------------------
# Monthly / seasonal transform
# ---------------------------------------------------------------------------


def _transform_monthly(ws) -> None:
    """
    Transform the 'bulletin' sheet.

    Before (cols A-J = 10, two sections):
      Rows 1-2: title merges A1:J1, A2:J2
      Row 3:  headers (A=РЕКА, B=ПУНКТ, C-J)
      Row 4:  {{HEADER.BASIN_NAME}} merged A4:J4
      Row 5:  {{DATA.*}} placeholders
      Rows 7-8: reservoir section titles merged A7:J7, A8:J8
      Row 9:  reservoir headers (A=РЕКА, B=ПУНКТ, ...)
      Row 10: reservoir {{DATA.*}} placeholders

    After (cols A-L = 12):
      Rows 1-2: title merges A1:L1, A2:L2
      Row 3:  A3=№, B3=Бассейн, C3 onward = old A3 onward
      Row 4:  separator A4:L4, NO {{HEADER.*}} content
      Row 5:  A5={{DATA.BASIN_NO}}, B5={{DATA.BASIN_NAME}}, C5 onward = old A5
      Rows 7-8: reservoir titles A7:L7, A8:L8
      Row 9:  A9=№, B9=Бассейн, C9 onward = old A9 onward
      Row 10: A10={{DATA.BASIN_NO}}, B10={{DATA.BASIN_NAME}}, C10 = old A10
    """
    new_max_col = 12  # L

    # 1. Capture style references BEFORE insert
    ref_hdr = ws["A3"]   # "РЕКА" — style source for №/Бассейн in header rows
    ref_hdr_font = copy.copy(ref_hdr.font)
    ref_hdr_fill = copy.copy(ref_hdr.fill)
    ref_hdr_border = copy.copy(ref_hdr.border)
    ref_hdr_number_format = ref_hdr.number_format

    ref_data = ws["A5"]  # "{{DATA.RIVER_NAME}}" — style source for data cells
    ref_data_font = copy.copy(ref_data.font)
    ref_data_fill = copy.copy(ref_data.fill)
    ref_data_border = copy.copy(ref_data.border)
    ref_data_number_format = ref_data.number_format

    # 2. Snapshot merges, unmerge all, insert, rebuild
    snapshots = _snapshot_merges(ws)
    _unmerge_all(ws)
    ws.insert_cols(1, N_INSERT)
    # Now old col N → new col N+2; A and B are empty
    _rebuild_merges(ws, snapshots, delta=N_INSERT)

    # 3. Extend title merges rows 1, 2, 7, 8 to full new width
    _set_title_merge_width(ws, 1, new_max_col)
    _set_title_merge_width(ws, 2, new_max_col)
    _set_title_merge_width(ws, 7, new_max_col)
    _set_title_merge_width(ws, 8, new_max_col)

    # 4. Row 4: HEADER.BASIN_NAME banner → clear and extend to A4:L4
    #    After shift, old A4 content is now C4; old merge A4:J4 → C4:L4
    ws["C4"].value = None
    ws["A4"].value = None
    _set_title_merge_width(ws, 4, new_max_col)

    # 5. MAIN block header row 3: A3=№, B3=Бассейн
    ws["A3"].value = "№"
    ws["A3"].font = ref_hdr_font
    ws["A3"].fill = ref_hdr_fill
    ws["A3"].border = ref_hdr_border
    ws["A3"].number_format = ref_hdr_number_format
    ws["A3"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws["B3"].value = "Бассейн"
    ws["B3"].font = ref_hdr_font
    ws["B3"].fill = ref_hdr_fill
    ws["B3"].border = ref_hdr_border
    ws["B3"].number_format = ref_hdr_number_format
    ws["B3"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 6. MAIN block data row 5: A5={{DATA.BASIN_NO}}, B5={{DATA.BASIN_NAME}}
    ws["A5"].value = "{{DATA.BASIN_NO}}"
    ws["A5"].font = ref_data_font
    ws["A5"].fill = ref_data_fill
    ws["A5"].border = ref_data_border
    ws["A5"].number_format = ref_data_number_format
    ws["A5"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws["B5"].value = "{{DATA.BASIN_NAME}}"
    ws["B5"].font = ref_data_font
    ws["B5"].fill = ref_data_fill
    ws["B5"].border = ref_data_border
    ws["B5"].number_format = ref_data_number_format
    ws["B5"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 7. RESERVOIR block header row 9: A9=№, B9=Бассейн
    ws["A9"].value = "№"
    ws["A9"].font = ref_hdr_font
    ws["A9"].fill = ref_hdr_fill
    ws["A9"].border = ref_hdr_border
    ws["A9"].number_format = ref_hdr_number_format
    ws["A9"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws["B9"].value = "Бассейн"
    ws["B9"].font = ref_hdr_font
    ws["B9"].fill = ref_hdr_fill
    ws["B9"].border = ref_hdr_border
    ws["B9"].number_format = ref_hdr_number_format
    ws["B9"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 8. RESERVOIR block data row 10: A10={{DATA.BASIN_NO}}, B10={{DATA.BASIN_NAME}}
    ws["A10"].value = "{{DATA.BASIN_NO}}"
    ws["A10"].font = ref_data_font
    ws["A10"].fill = ref_data_fill
    ws["A10"].border = ref_data_border
    ws["A10"].number_format = ref_data_number_format
    ws["A10"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws["B10"].value = "{{DATA.BASIN_NAME}}"
    ws["B10"].font = ref_data_font
    ws["B10"].fill = ref_data_fill
    ws["B10"].border = ref_data_border
    ws["B10"].number_format = ref_data_number_format
    ws["B10"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 9. Set column widths
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 18.0


# ---------------------------------------------------------------------------
# Verification helper
# ---------------------------------------------------------------------------


def _verify(ws, sheet_type: str) -> dict:
    """
    Verify post-transform correctness. Returns a dict with boolean results.

    sheet_type: 'pentadal' | 'monthly'
    """
    if sheet_type == "pentadal":
        header_row = 5
        data_row = 8
        basin_data_tag = "{{DATA.BASIN_RU}}"
    else:
        header_row = 3
        data_row = 5
        basin_data_tag = "{{DATA.BASIN_NAME}}"

    # (a) № and Бассейн headers present
    no_header_present = ws.cell(header_row, 1).value == "№"
    basin_header_present = ws.cell(header_row, 2).value == "Бассейн"

    # (b) DATA tags at A/B of data row
    basin_no_data_tag = ws.cell(data_row, 1).value == "{{DATA.BASIN_NO}}"
    basin_name_data_tag = ws.cell(data_row, 2).value == basin_data_tag

    # (c) No {{HEADER.* anywhere
    header_violations = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{HEADER." in cell.value:
                header_violations.append(
                    f"{cell.coordinate}: {cell.value!r}"
                )

    return {
        "no_header_present": no_header_present,
        "basin_header_present": basin_header_present,
        "basin_no_data_tag": basin_no_data_tag,
        "basin_name_data_tag": basin_name_data_tag,
        "no_header_tags": len(header_violations) == 0,
        "header_tag_violations": header_violations,
    }


def _collect_dump(ws) -> dict:
    """Collect non-empty cells, merged ranges, and column widths."""
    cells = [
        (cell.coordinate, cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    merges = sorted(str(mr) for mr in ws.merged_cells.ranges)
    col_widths = {
        col: cd.width
        for col, cd in ws.column_dimensions.items()
        if cd.width
    }
    return {"cells": cells, "merges": merges, "col_widths": col_widths}


# ---------------------------------------------------------------------------
# Per-file orchestrator
# ---------------------------------------------------------------------------


def transform_file(path: Path) -> dict:
    """
    Transform one template file in-place.

    Returns a status dict suitable for reporting.
    """
    fname = path.name
    result: dict = {"file": str(path), "status": None, "reason": ""}

    # Warn about Excel lock files
    lock = path.parent / f"~${fname}"
    if lock.exists():
        print(
            f"  WARNING: lock file {lock.name} exists — "
            "file may be open in Excel"
        )

    is_pentadal = "pentadal" in fname or "decadal" in fname
    is_monthly = "monthly" in fname or "seasonal" in fname

    if not (is_pentadal or is_monthly):
        result["status"] = "error"
        result["reason"] = "Cannot determine template type from filename"
        return result

    try:
        wb = openpyxl.load_workbook(str(path))
    except Exception as exc:
        result["status"] = "error"
        result["reason"] = f"Cannot open workbook: {exc}"
        return result

    sheet_name = "1 пентада" if is_pentadal else "bulletin"
    if sheet_name not in wb.sheetnames:
        result["status"] = "error"
        result["reason"] = f"Sheet '{sheet_name}' not found"
        return result

    ws = wb[sheet_name]

    # Idempotency check
    idempotency_row = 5 if is_pentadal else 3
    if ws.cell(idempotency_row, 1).value == "№":
        result["status"] = "skipped"
        result["reason"] = (
            f"Row {idempotency_row} col A already contains '№' "
            "— already transformed"
        )
        result.update(_collect_dump(ws))
        return result

    # Backup (skip if .bak already exists to preserve the original)
    bak_path = path.parent / (fname + ".bak")
    if not bak_path.exists():
        import shutil
        shutil.copy2(str(path), str(bak_path))
        print(f"  Backed up to: {bak_path.name}")
    else:
        print(f"  Backup exists: {bak_path.name} — skipping backup")

    # Apply transform
    try:
        if is_pentadal:
            _transform_pentadal(ws)
            sheet_type = "pentadal"
        else:
            _transform_monthly(ws)
            sheet_type = "monthly"
    except Exception as exc:
        import traceback
        result["status"] = "error"
        result["reason"] = f"Transform failed: {exc}"
        result["traceback"] = traceback.format_exc()
        return result

    # Save
    try:
        wb.save(str(path))
    except Exception as exc:
        result["status"] = "error"
        result["reason"] = f"Save failed: {exc}"
        return result

    # Reload for verification (fresh parse)
    wb2 = openpyxl.load_workbook(str(path))
    ws2 = wb2[sheet_name]

    result["verification"] = _verify(ws2, sheet_type)
    result.update(_collect_dump(ws2))
    result["status"] = "transformed"
    result["reason"] = "OK"
    result["bak_exists"] = bak_path.exists()
    return result


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Add №/Бассейн leading columns to Tajik bulletin Excel templates."
        )
    )
    parser.add_argument(
        "--templates-dir",
        type=Path,
        default=_DEFAULT_TMPL_DIR,
        help="Path to taj_data_forecast_tools/templates/ directory",
    )
    args = parser.parse_args()

    tmpl_dir: Path = args.templates_dir
    if not tmpl_dir.exists():
        print(f"ERROR: templates directory not found: {tmpl_dir}")
        sys.exit(1)

    print(f"Templates directory: {tmpl_dir}")
    print()

    all_results: dict = {}
    for fname in TJ_TEMPLATES:
        path = tmpl_dir / fname
        print(f"Processing: {fname}")
        if not path.exists():
            print("  SKIPPED: file not found")
            all_results[fname] = {"status": "not_found", "file": str(path)}
            print()
            continue

        result = transform_file(path)
        all_results[fname] = result
        print(f"  Status: {result['status']} — {result.get('reason', '')}")

        if "traceback" in result:
            print(result["traceback"])

        if "verification" in result:
            v = result["verification"]
            print("  Checks:")
            print(f"    (a) № header present:        {v['no_header_present']}")
            print(f"    (a) Бассейн header present:   {v['basin_header_present']}")
            print(f"    (b) BASIN_NO data tag A/B:    {v['basin_no_data_tag']}")
            print(f"    (b) basin name data tag:      {v['basin_name_data_tag']}")
            print(f"    (c) Zero {{HEADER.* tags:      {v['no_header_tags']}")
            if v["header_tag_violations"]:
                print(f"    VIOLATIONS: {v['header_tag_violations']}")

        if "cells" in result:
            print("  Non-empty cells:")
            for coord, val in result["cells"]:
                print(f"    {coord}: {val!r}")
            print("  Merged ranges:")
            for mr in result["merges"]:
                print(f"    {mr}")
            print("  Column widths:")
            for col, w in sorted(result["col_widths"].items()):
                print(f"    {col}: {w}")

        if "bak_exists" in result:
            print(f"  (d) .bak backup exists:       {result['bak_exists']}")
        print()

    # Summary table
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for fname, r in all_results.items():
        v = r.get("verification", {})
        checks = [
            v.get("no_header_present"),
            v.get("basin_header_present"),
            v.get("basin_no_data_tag"),
            v.get("basin_name_data_tag"),
            v.get("no_header_tags"),
        ]
        checks_str = " ".join(
            "OK" if c is True else ("FAIL" if c is False else "?")
            for c in checks
        )
        print(f"  {fname}: {r['status']} | {checks_str}")


if __name__ == "__main__":
    main()
