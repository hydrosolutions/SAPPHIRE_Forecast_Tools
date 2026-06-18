import os
import math
from copy import copy
import openpyxl
import panel as pn
import pandas as pd
from typing import List

# ieassyreport
from ieasyreports.settings import TagSettings, ReportGeneratorSettings
from ieasyreports.core.tags.tag import Tag
from ieasyreports.core.report_generator import DefaultReportGenerator

# Logging
import logging

# Configure logging
logger = logging.getLogger(__name__)

# Local imports
from .reports import SapphireReport
import tag_library as tl


# Custom class
# Overwrite the default report generator to add functinality to write to
# specific sheets
class MultiSheetReportGenerator(DefaultReportGenerator):
    def __init__(
        self,
        tags: List[Tag],
        template: str,
        templates_directory_path: str,
        reports_directory_path: str,
        tag_settings: TagSettings,
        requires_header: bool = False,
        sheet: int = 0
    ):
        self.tags = {tag.name: tag for tag in tags}
        self.template_filename = template
        self.templates_directory_path = templates_directory_path
        self.reports_directory_path = reports_directory_path
        self.template = self.open_template_file()
        self.tag_settings = tag_settings
        self.sheet = self.template.worksheets[sheet]

        self.validated = False

        self.requires_header_tag = requires_header
        self.header_tag_info = {}
        self.data_tags_info = []
        self.general_tags = {}

class MultiSectionReportGenerator(DefaultReportGenerator):
    """Report generator that supports multiple DATA rows in one sheet.

    The template may have DATA tags on multiple rows (one per section).
    Each section is processed independently; only one section may carry a
    HEADER tag.  Sections are identified by the row number of their DATA tags.

    Args:
        tags_per_section: Optional list of per-section tag lists.  When
            supplied, each section's DATA tags are resolved from the
            corresponding list rather than from ``tags``.  Tag matching is
            by name so that multiple sections can share tag names with
            different ``get_value_fn`` implementations.
    """

    def __init__(
        self,
        tags: List[Tag],
        template: str,
        templates_directory_path: str,
        reports_directory_path: str,
        tag_settings: TagSettings,
        requires_header: bool = False,
        tags_per_section: "list[list[Tag]] | None" = None,
    ):
        super().__init__(
            tags=tags,
            template=template,
            templates_directory_path=templates_directory_path,
            reports_directory_path=reports_directory_path,
            tag_settings=tag_settings,
            requires_header=requires_header,
        )
        # per-section tag lookup: list[dict[name -> Tag]]
        self._tags_per_section: "list[dict[str, Tag]] | None" = None
        if tags_per_section is not None:
            self._tags_per_section = [
                {t.name: t for t in section_tags}
                for section_tags in tags_per_section
            ]
            # Seed "special" context so per-section tags produce the
            # correct {{DATA.X}} / {{HEADER.X}} form when tag.replace
            # looks them up by name during substitution.
            for section_tags in tags_per_section:
                for t in section_tags:
                    if t.data:
                        t.set_context({"special": tag_settings.data_tag})
                    elif t.header:
                        t.set_context({"special": tag_settings.header_tag})
        # sections: ordered list of {"data_row": int, "cells": [{"tag": Tag, "cell": Cell}]}
        self.sections: "list[dict]" = []
        self._header_section_idx: "int | None" = None

    # ------------------------------------------------------------------
    # Validation overrides
    # ------------------------------------------------------------------

    def _check_template_tags(self) -> None:
        """Scan template cells; group DATA tags by row into self.sections."""
        sections_by_row: "dict[int, list[dict]]" = {}
        for cell in self.iter_cells():
            if cell.value is None:
                continue
            for tag in self._parse_template_tag(cell.value):
                tag_info = self._decode_template_tag(tag)
                if tag_info["tag"] not in self.tags:
                    from ieasyreports.exceptions import InvalidTagException
                    raise InvalidTagException(
                        f"The following tag is not supported: {tag_info['tag']}"
                    )
                tag_obj = self.tags[tag_info["tag"]]
                if tag_info["tag_type"] == self.tag_settings.header_tag:
                    if self.header_tag_info:
                        from ieasyreports.exceptions import MultipleHeaderTagsException
                        raise MultipleHeaderTagsException("Multiple header tags found.")
                    tag_obj.set_context({"special": self.tag_settings.header_tag})
                    self.header_tag_info["tag"] = tag_obj
                    self.header_tag_info["cell"] = cell
                    self.header_tag_info["row"] = cell.row
                elif tag_info["tag_type"] == self.tag_settings.data_tag:
                    tag_obj.set_context({"special": self.tag_settings.data_tag})
                    row = cell.row
                    if row not in sections_by_row:
                        sections_by_row[row] = []
                    sections_by_row[row].append({"tag": tag_obj, "cell": cell})
                else:
                    if tag_obj not in self.general_tags:
                        self.general_tags[tag_obj] = []
                    self.general_tags[tag_obj].append(cell)

        # Build ordered sections list
        data_rows = sorted(sections_by_row.keys())
        for idx, row in enumerate(data_rows):
            self.sections.append({"data_row": row, "cells": sections_by_row[row]})

        # Populate legacy data_tags_info for compatibility with parent helpers
        if self.sections:
            self.data_tags_info = self.sections[0]["cells"]

        # Identify which section contains the header tag
        if self.header_tag_info:
            header_row = self.header_tag_info.get("row")
            for idx, section in enumerate(self.sections):
                if section["data_row"] == header_row + 1:
                    self._header_section_idx = idx
                    break

    def _validate_data_tags(self) -> None:
        """Allow multiple data rows; each section's tags must be in the same row."""
        if not self.sections:
            from ieasyreports.exceptions import MissingDataTagException
            raise MissingDataTagException("At least one DATA tag is required.")
        for section in self.sections:
            rows = {info["cell"].row for info in section["cells"]}
            if len(rows) > 1:
                from ieasyreports.exceptions import InvalidTagException
                raise InvalidTagException(
                    f"All DATA tags within a section must be in the same row, got rows {rows}."
                )

    # ------------------------------------------------------------------
    # Formula helpers
    # ------------------------------------------------------------------

    def _write_perc_formulas(self, row, section_cells):
        """Write the K (% of norm) and L (% of previous year) Excel formulas
        for one data row, deriving column letters from the section's tags.

        K = midpoint of the forecast interval (Q_MIN..Q_MAX) as % of NORM.
        L = the same midpoint as % of the previous year's discharge (VNORM).
        IFERROR yields a blank cell when a denominator is empty/zero. Overwrites
        whatever the PERC_NORM / PERC_PREVYEAR tags wrote into K / L.
        """
        from openpyxl.utils import get_column_letter
        cols = {
            ci["tag"].name: get_column_letter(ci["cell"].column)
            for ci in section_cells
        }
        qmin = cols.get("Q_MIN")
        qmax = cols.get("Q_MAX")
        norm = cols.get("NORM")
        vnorm = cols.get("VNORM")
        k_col = cols.get("PERC_NORM")
        l_col = cols.get("PERC_PREVYEAR")
        if k_col and qmin and qmax and norm:
            self.sheet[f"{k_col}{row}"] = (
                f'=IFERROR(ROUND(({qmin}{row}+{qmax}{row})/2/{norm}{row}*100,0),"")'
            )
        if l_col and qmin and qmax and vnorm:
            self.sheet[f"{l_col}{row}"] = (
                f'=IFERROR(ROUND(({qmin}{row}+{qmax}{row})/2/{vnorm}{row}*100,0),"")'
            )

    def _numerify_value_cells(self, row, section_cells):
        """Convert the value columns (Q_MIN/Q_MAX/V_MIN/V_MAX/NORM/VNORM) of one
        data row from the comma-decimal text the tags wrote into real numeric
        cells, preserving the displayed precision via a number format.

        The tag-written string already carries the correct rounding (e.g.
        '12,3'); we parse it back to a number and set a number format matching
        its decimal count ('0', '0.0', '0.00'), so Excel shows the same value
        (with the locale decimal separator) but stores a number. Blank / dash /
        unparseable cells are left untouched.
        """
        from openpyxl.utils import get_column_letter
        value_tags = {"Q_MIN", "Q_MAX", "V_MIN", "V_MAX", "NORM", "VNORM", "Q_LAST_YEAR"}
        for ci in section_cells:
            if ci["tag"].name not in value_tags:
                continue
            col = get_column_letter(ci["cell"].column)
            cell = self.sheet[f"{col}{row}"]
            raw = cell.value
            if not isinstance(raw, str):
                continue
            text = raw.strip()
            if not text:
                continue
            normalized = text.replace(",", ".")
            try:
                num = float(normalized)
            except ValueError:
                continue
            decimals = len(text.split(",", 1)[1]) if "," in text else 0
            if decimals == 0:
                cell.value = int(round(num))
                cell.number_format = "0"
            else:
                cell.value = num
                cell.number_format = "0." + "0" * decimals

    # ------------------------------------------------------------------
    # Header-bearing section rendering
    # ------------------------------------------------------------------

    def _handle_header_and_data_tags(self, grouped_data):
        original_header_cell = self.header_tag_info["cell"]
        original_header_row = original_header_cell.row
        original_header_col = original_header_cell.col_idx
        current_row = original_header_row
        for header_value, item_group in grouped_data.items():
            cell = self.sheet.cell(
                row=current_row,
                column=original_header_col
            )
            cell.value = header_value
            current_row += 1
            for item in item_group:
                for idx, data_tag in enumerate(self.data_tags_info):
                    tag = data_tag["tag"]
                    tag.set_context({"obj": item})
                    data_cell = self.sheet.cell(row=current_row, column=data_tag["cell"].column)
                    data_cell.value = tag.replace(data_cell.value)
                self._numerify_value_cells(current_row, self.data_tags_info)
                self._write_perc_formulas(current_row, self.data_tags_info)
                current_row += 1

    # ------------------------------------------------------------------
    # Multi-section report generation
    # ------------------------------------------------------------------

    def generate_report_multi(
        self,
        list_objects_per_section: "list[list | None]",
        output_path: "str | None" = None,
        output_filename: "str | None" = None,
    ) -> None:
        """Render each section with its own list of objects.

        Args:
            list_objects_per_section: One entry per section (ordered by
                ascending data-row).  ``None`` or an empty list causes the
                section rows to be deleted.
            output_path: Directory for the output file.
            output_filename: File name for the output file.
        """
        if not self.validated:
            from ieasyreports.exceptions import TemplateNotValidatedException
            raise TemplateNotValidatedException(
                "Template must be validated first."
            )

        # Determine section bounds (start_row, end_row) — used for deletion.
        # Template layout: section i spans from (data_rows[i-1]+2) to (data_rows[i]+1).
        # For the first section it starts at row 1.
        data_rows = [s["data_row"] for s in self.sections]
        section_bounds = []
        for i, dr in enumerate(data_rows):
            start = (data_rows[i - 1] + 2) if i > 0 else 1
            end = dr + 1  # include the one blank separator row after data row
            section_bounds.append((start, end))

        # Process sections in reverse order so row shifts don't affect earlier sections
        for idx in range(len(self.sections) - 1, -1, -1):
            section = self.sections[idx]
            objects = list_objects_per_section[idx] if idx < len(list_objects_per_section) else None

            if not objects:
                # Delete this section's rows.
                start_row, end_row = section_bounds[idx]
                row_count = end_row - start_row + 1
                # openpyxl's delete_rows does NOT drop merged ranges in the
                # deleted rows; left stale, a later _insert_rows -> unmerge_cells
                # raises KeyError on the missing cells. Unmerge them first (while
                # the cells still exist).
                for mr in list(self.sheet.merged_cells.ranges):
                    if mr.min_row >= start_row and mr.max_row <= end_row:
                        self.sheet.unmerge_cells(str(mr))
                self.sheet.delete_rows(start_row, row_count)
                continue

            data_row = section["data_row"]
            section_cells = section["cells"]

            # Resolve tags for this section
            if self._tags_per_section and idx < len(self._tags_per_section):
                section_tag_map = self._tags_per_section[idx]
            else:
                section_tag_map = None

            if idx == self._header_section_idx:
                # Header-bearing section: use parent group-by-header logic
                # temporarily swap data_tags_info to this section's cells
                saved_data_tags_info = self.data_tags_info
                self.data_tags_info = section_cells
                if section_tag_map:
                    for cell_info in self.data_tags_info:
                        name = cell_info["tag"].name
                        if name in section_tag_map:
                            cell_info["tag"] = section_tag_map[name]
                grouped = self._create_header_grouping(objects)
                self._prepare_structure(grouped)
                self._handle_header_and_data_tags(grouped)
                self.data_tags_info = saved_data_tags_info
            else:
                # Header-less section: flat list rendering.
                # Replicate the data-row template (values + styles + merged cells)
                # into each new row via _copy_cell_range — mirrors _prepare_structure.
                n_extra = len(objects) - 1
                if n_extra > 0:
                    self._insert_rows(data_row, n_extra)
                    dest_ranges = [(data_row + i, 1) for i in range(1, n_extra + 1)]
                    self._copy_cell_range(
                        (data_row, 1),
                        (data_row, 25),
                        dest_ranges,
                    )

                current_row = data_row
                for obj in objects:
                    for cell_info in section_cells:
                        tag = cell_info["tag"]
                        if section_tag_map and tag.name in section_tag_map:
                            tag = section_tag_map[tag.name]
                        tag.set_context({"obj": obj})
                        dest_cell = self.sheet.cell(
                            row=current_row, column=cell_info["cell"].column
                        )
                        dest_cell.value = tag.replace(dest_cell.value)
                    self._numerify_value_cells(current_row, section_cells)
                    self._write_perc_formulas(current_row, section_cells)
                    current_row += 1

        self._handle_general_tags()
        self.save_report(output_filename, output_path)


def round_percentage_to_integer_string(value: float) -> int:
    '''
    Round percentage to integers.

    Args:
        value (float): The percentage value to round.

    Returns:
        str: The rounded percentage value. An empty string is returned in case of
            a negative input value.
    '''
    try:
        # Test if value is NaN
        if math.isnan(value):
            return None
        
        if not isinstance(value, float):
            raise TypeError('Input value must be a float')

        if value < 0.0:
            return None
        return f'{round(value)}'
    except TypeError as e:
        print(f'Error in round_percentage: {e}')
        return None
    except Exception as e:
        print(f'Error in round_percentage: {e}')
        return None

# def round_percentage_to_comma_separated_string(value: float) -> str:
#     '''
#     Round percentage to 0 decimals for values ge  100, to 1 decimal for values
#     ge  10 and to 2 decimals for values ge  0.

#     Args:
#         value (str): The percentage value to round.

#     Returns:
#         str: The rounded percentage value. An empty string is returned in case of
#             a negative input value.
#     '''
#     try:
#         if not isinstance(value, float):
#             raise TypeError('Input value must be a float')

#         if math.isclose(value, 100.0):
#             string = "100"
#         elif abs(value) > 0.0 and abs(value) < 10.0:
#             string = "{:.2f}".format(round(value, 2))
#         elif abs(value) >= 10.0 and abs(value) < 100.0:
#             string = "{:.1f}".format(round(value, 1))
#         else:
#             string = "{:.0f}".format(round(value, 0))
#         # Replace . in string with ,
#         string = string.replace('.', ',')
#         return string
#     except TypeError as e:
#         print(f'Error in round_percentage: {e}')
#         return None
#     except Exception as e:
#         print(f'Error in round_percentage: {e}')
#         return None

def round_discharge_to_comma_separated_string(value: float) -> str:
    '''
    Round discharge to 0 decimals for values ge 100, to 1 decimal for values
    ge 10 and to 2 decimals for values ge 0.

    Args:
        value (str): The discharge value to round.

    Returns:
        str: The rounded discharge value. An empty string is returned in case of
            a negative input value.

    Examples:
        >>> round_discharge(0.0)
        '0'
        >>> round_discharge(0.123)
        '0.1'
        >>> round_discharge(0.0123)
        '0.01'
        >>> round_discharge(0.00623)
        '0.01'
        >>> round_discharge(1.0)
        '1'
        >>> round_discharge(1.23)
        '1.2'
        >>> round_discharge(1.0123)
        '1.01'
        >>> round_discharge(10.123)
        '10.1'
        >>> round_discharge(100.123)
        '100'
        >>> round_discharge(1000.123)
        '1000'
    '''
    try:
        if not isinstance(value, float):
            raise TypeError('Input value must be a float')

        # Return an empty string if the input value is negative
        if value < 0.0:
            string = " "
        # Test if the input value is close to zero
        elif math.isclose(value, 0.0):
            string = "0"
        elif value > 0.0 and value < 10.0:
            string = "{:.2f}".format(round(value, 2))
        elif value >= 10.0 and value < 100.0:
            string = "{:.1f}".format(round(value, 1))
        else:
            string = "{:.0f}".format(round(value, 0))
        # Replace . in string with ,
        string = string.replace('.', ',')
        return string
    except TypeError as e:
        print(f'Error in round_discharge: {e}')
        return None
    except Exception as e:
        print(f'Error in round_discharge: {e}')
        return None

def copy_worksheet(report_settings, temp_bulletin_file_name, bulletin_file_name,
                   header_df, horizon):
    """
    Copy the sheet 1 of the generated report to the appropriate sheet in the final bulletin.

    Args:
        report_settings (ReportGeneratorSettings): The report settings.
        temp_bulletin_file_name (str): The temporary bulletin file name.
        bulletin_file_name (str): The bulletin file name.
        header_df (pandas.DataFrame): A DataFrame containing the header information for the Excel file.

    Returns:
        None
    """
    sapphire_forecast_horizon = horizon
    if sapphire_forecast_horizon == 'pentad':
        horizon_string_ru = "пентада"
    elif sapphire_forecast_horizon == 'decad':
        horizon_string_ru = "декада"
    elif sapphire_forecast_horizon == 'month':
        horizon_string_ru = "месяц"
    else:
        raise ValueError(f"Invalid sapphire_forecast_horizon: {sapphire_forecast_horizon}")

    def _sheet_key(hdf, horizon):
        if horizon == 'pentad':
            return int(hdf['pentad'].values[0])
        elif horizon == 'decad':
            return int(hdf['decad'].values[0])
        elif horizon == 'month':
            return int(hdf['month_number'].values[0])
        return 0
    
    # Now copy the sheet 1 of the generated report to the appropriate sheet in the final bulletin
    # Load the generated report
    try:
        generated_report = openpyxl.load_workbook(
            os.path.join(
                report_settings.report_output_path, temp_bulletin_file_name))
    except Exception as e:
        raise Exception(f"Error loading the generated report: {e}")

    final_path = os.path.join(report_settings.report_output_path, bulletin_file_name)
    temp_path = os.path.join(report_settings.report_output_path, temp_bulletin_file_name)
    sheet_title = f"{_sheet_key(header_df, sapphire_forecast_horizon)} {horizon_string_ru}"

    # If the final bulletin file exists, try to load it and merge the new sheet in.
    # If it is corrupt or missing, fall through to the rename path.
    final_bulletin = None
    if os.path.exists(final_path):
        try:
            final_bulletin = openpyxl.load_workbook(final_path)
        except Exception as e:
            print(f"DEBUG: copy_worksheet: Final bulletin corrupt or unreadable ({e}). "
                  f"Deleting and recreating from temp.")
            os.remove(final_path)
            final_bulletin = None

    if final_bulletin is not None:
        print(f"DEBUG: write_to_excel: initial final_bulletin.sheetnames: {final_bulletin.sheetnames}")

        # Remove pre-existing sheet for this period if present
        if sheet_title in final_bulletin.sheetnames:
            print(f"DEBUG: write_to_excel: Removing sheet for {sapphire_forecast_horizon} "
                  f"{_sheet_key(header_df, sapphire_forecast_horizon)}")
            final_bulletin.remove(final_bulletin[sheet_title])

        # Create a new sheet in the destination workbook and copy content cell-by-cell.
        # This avoids the unsafe _parent/_add_sheet hack that leaves dangling style indices.
        src = generated_report.active
        new_ws = final_bulletin.create_sheet(title=sheet_title)

        for row in src.iter_rows():
            for cell in row:
                nc = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    nc.font = copy(cell.font)
                    nc.border = copy(cell.border)
                    nc.fill = copy(cell.fill)
                    nc.alignment = copy(cell.alignment)
                    nc.number_format = cell.number_format
                    nc.protection = copy(cell.protection)

        for mr in list(src.merged_cells.ranges):
            new_ws.merge_cells(str(mr))

        for col_letter, dim in src.column_dimensions.items():
            if dim.width is not None:
                new_ws.column_dimensions[col_letter].width = dim.width

        for row_idx, dim in src.row_dimensions.items():
            if dim.height is not None:
                new_ws.row_dimensions[row_idx].height = dim.height

        # Save the final bulletin
        final_bulletin.save(final_path)

        # Close the workbooks
        generated_report.close()
        final_bulletin.close()

        # Delete the generated report temp file
        os.remove(temp_path)

    else:
        # Final bulletin does not exist (or was corrupt and deleted above):
        # rename the temp file into place as the new final.
        os.rename(temp_path, final_path)
        # Load to rename the sheet correctly, then save.
        final_bulletin = openpyxl.load_workbook(final_path)
        final_bulletin.active.title = sheet_title
        final_bulletin.save(final_path)
        final_bulletin.close()

def oder_sites_list_according_to_bulletin_order(sites_list):
        """Order the sites_list according to the order in the attribute bulletin_order of each site"""
        # Get the basin and bulletin order for each site
        df = pd.DataFrame({
            'codes': [site.code for site in sites_list],
            'basins': [site.basin_ru for site in sites_list],
            'bulletin_order': [site.bulletin_order for site in sites_list]
        })
        # Sort the sites_list according to the basin and bulletin order
        df = df.sort_values(by=['basins', 'bulletin_order'])
        print(f"Ordered sites: {df}")
        # Get the ordered list of codes
        ordered_codes = df['codes'].tolist()
        # Iterate over the ordered_codes and add sites in sites_list to 
        # ordered_sites_list in the order of ordered_codes
        ordered_sites_list = []
        for code in ordered_codes:
            for site in sites_list:
                if site.code == code:
                    ordered_sites_list.append(site)
        return ordered_sites_list

# Function to write data to Excel
def write_to_excel(sites_list, bulletin_sites, header_df, env_file_path,
                   tag_settings=None, horizon=None):
    """
    Writes data to an Excel file.

    Args:
        sites_list (list): A list of sites for which to write data to the Excel file.
        bulletin_sites (list): A list of sites for which to write data to the Excel file.
        header_df (pandas.DataFrame): A DataFrame containing the header information for the Excel file.
        env_file_path (str): The path to the environment file.
        tag_settings (TagSettings): The tag settings.

    Returns:
        None

    """

    # Show the loading spinner
    #indicator.value = True

    # Get the forecast horizon from the caller (widget value)
    sapphire_forecast_horizon = horizon
    if sapphire_forecast_horizon is None:
        raise ValueError("horizon parameter is required")
    if sapphire_forecast_horizon not in ['pentad', 'decad', 'month', 'season']:
        raise ValueError(f"horizon must be 'pentad', 'decad', 'month', or 'season', got '{sapphire_forecast_horizon}'")
    print(f"DEBUG: write_to_excel: sapphire_forecast_horizon: {sapphire_forecast_horizon}")


    print('DEBUG: write_to_excel: Initializing report generator ...')

    # Define tag & report settings
    tag_settings = TagSettings() if tag_settings is None else tag_settings
    report = SapphireReport(name="Test report", env_file_path=env_file_path)
    report_settings = report.define_settings(env_file_path)

    # Define Tags
    # region tags
    # Some tags are only defined for certain forecast horizons
    if sapphire_forecast_horizon == 'pentad':
        pentad_tag = Tag(
            name='PENTAD',
            get_value_fn=header_df['pentad'].values[0],
            tag_settings=tag_settings)
        
        day_start_pentad_tag = Tag(
            name='DAY_START_PENTAD',
            get_value_fn=header_df['day_start_pentad'].values[0],
            tag_settings=tag_settings)

        day_end_pentad_tag = Tag(
            name='DAY_END_PENTAD',
            get_value_fn=header_df['day_end_pentad'].values[0],
            tag_settings=tag_settings)

    elif sapphire_forecast_horizon == 'decad':
        decad_tag = Tag(
            name='DEKAD',
            get_value_fn=header_df['decad'].values[0],
            tag_settings=tag_settings)

        day_start_decad_tag = Tag(
            name='DAY_START_DEKAD',
            get_value_fn=header_df['day_start_decad'].values[0],
            tag_settings=tag_settings)

        day_end_decad_tag = Tag(
            name='DAY_END_DEKAD',
            get_value_fn=header_df['day_end_decad'].values[0],
            tag_settings=tag_settings)

    if sapphire_forecast_horizon in ('month', 'season'):
        # Tag names must match the last segment of the template tag
        # (ieasyreports splits on '.' and uses only the suffix).
        # Prefixes like HEADER./DATA./ in the template route the tag to
        # the header or data category; general tags have no prefix.
        def _as_float(v):
            """Convert numpy/pandas scalars to Python float; return None on failure."""
            if v is None:
                return None
            try:
                import math as _math
                fv = float(v)
                if _math.isnan(fv):
                    return None
                return fv
            except (TypeError, ValueError):
                return None

        def _fmt_discharge(v):
            """Render a float as comma-separated string, or blank for None/NaN."""
            fv = _as_float(v)
            if fv is None:
                return ''
            return round_discharge_to_comma_separated_string(fv) or ''

        def _fmt_percentage(v):
            """Render a percentage as int string, or blank for None/NaN."""
            fv = _as_float(v)
            if fv is None:
                return ''
            return round_percentage_to_integer_string(fv) or ''
        fc_month_tag = Tag(
            name='MONTH',
            get_value_fn=header_df['month_str_nom_ru'].values[0],
            tag_settings=tag_settings,
        )
        fc_year_tag = Tag(
            name='YEAR',
            get_value_fn=header_df['year'].values[0],
            tag_settings=tag_settings,
        )
        fc_prevyear_tag = Tag(
            name='PrevYear',
            get_value_fn=header_df['prev_year'].values[0],
            tag_settings=tag_settings,
        )
        # Derive period labels: quarterly bounds for month, seasonal bounds for season.
        # All sites share the same period range, so the first matching site wins.
        _month_start_ru = header_df['month_str_nom_ru'].values[0]
        _month_end_ru = header_df['month_str_nom_ru'].values[0]
        for _site in bulletin_sites:
            if sapphire_forecast_horizon == 'month':
                vf = getattr(_site, 'quarterly_valid_from', None)
                vt = getattr(_site, 'quarterly_valid_to', None)
            else:
                vf = getattr(_site, 'seasonal_valid_from', None)
                vt = getattr(_site, 'seasonal_valid_to', None)
            if vf is not None and vt is not None:
                _month_start_ru = tl.get_month_str_case1(vf)
                _month_end_ru = tl.get_month_str_case1(vt)
                break
        fc_month_start_tag = Tag(
            name='MONTH_START',
            get_value_fn=_month_start_ru,
            tag_settings=tag_settings,
        )
        fc_month_end_tag = Tag(
            name='MONTH_END',
            get_value_fn=_month_end_ru,
            tag_settings=tag_settings,
        )

    month_string_nom_ru_tag = Tag(
        name='MONTH_STR_NOM_RU',
        get_value_fn=header_df['month_str_nom_ru'].values[0],
        tag_settings=tag_settings)

    month_string_gen_ru_tag = Tag(
        name='MONTH_STR_GEN_RU',
        get_value_fn=header_df['month_str_gen_ru'].values[0],
        tag_settings=tag_settings)

    year_tag = Tag(
        name='YEAR',
        get_value_fn=header_df['year'].values[0],
        tag_settings=tag_settings)

    header_tag = Tag(
            name='BASIN_RU',
            get_value_fn=lambda obj, **kwargs: obj.basin_ru,
            tag_settings=tag_settings,
            header=True)

    river_ru_tag = Tag(
            name='RIVER_NAME_RU',
            get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
            tag_settings=tag_settings,
            data=True)

    punkt_ru_tag = Tag(
        name='PUNKT_NAME_RU',
        get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
        tag_settings=tag_settings,
        data=True)

    model_tag = Tag(
        name='MODEL',
        get_value_fn=lambda obj, **kwargs: obj.forecast_model,
        tag_settings=tag_settings,
        data=True)

    linreg_predictor_tag = Tag(
        name='LINREG_PREDICTOR',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.linreg_predictor),
        tag_settings=tag_settings,
        data=True)

    forecast_tag = Tag(
        name='QEXP',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.forecast_expected),
        tag_settings=tag_settings,
        data=True
    )

    delta_tag = Tag(
        name='DELTA',
        get_value_fn=lambda obj, **kwargs: f"{round(obj.forecast_delta, 2)}".replace('.', ','),
        tag_settings=tag_settings,
        data=True
    )

    sdivsigma_tag = Tag(
        name='SDIVSIGMA',
        get_value_fn=lambda obj, **kwargs: f"{round(obj.forecast_sdivsigma, 2)}".replace('.', ','),
        tag_settings=tag_settings,
        data=True
    )

    forecast_lower_bound_tag = Tag(
        name='FORECAST_LOWER_BOUND',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.forecast_lower_bound),
        tag_settings=tag_settings,
        data=True
    )

    forecast_upper_bound_tag = Tag(
        name='FORECAST_UPPER_BOUND',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.forecast_upper_bound),
        tag_settings=tag_settings,
        data=True
    )

    dash_tag = Tag(
        name='DASH',
        get_value_fn='—',
        tag_settings=tag_settings,
        data=True
    )

    hydrograph_max_tag = Tag(
        name='HYDROGRAPH_MAX',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.hydrograph_max),
        tag_settings=tag_settings,
        data=True
    )

    hydrograph_min_tag = Tag(
        name='HYDROGRAPH_MIN',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.hydrograph_min),
        tag_settings=tag_settings,
        data=True
    )

    hydrograph_norm_tag = Tag(
        name='QNORM',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.hydrograph_norm),
        tag_settings=tag_settings,
        data=True
    )

    q_last_year_tag = Tag(
        name='Q_LAST_YEAR',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.last_year_q_pentad_mean),
        tag_settings=tag_settings,
        data=True
    )

    qdanger_tag = Tag(
        name='QDANGER',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.qdanger),
        tag_settings=tag_settings,
        data=True
    )

    perc_norm_tag = Tag(
        name='PERC_NORM',
        get_value_fn=lambda obj, **kwargs: round_percentage_to_integer_string(obj.perc_norm),
        tag_settings=tag_settings,
        data=True
    )

    def _safe_perc(numerator, denominator) -> str:
        """Compute numerator/denominator*100 as an integer string.

        Returns an empty string when either operand is None/NaN or the
        denominator is zero.  Uses the same formatter as perc_norm_tag
        (round_percentage_to_integer_string) for consistent display.
        """
        try:
            import math as _math
            n = float(numerator)
            d = float(denominator)
            if _math.isnan(n) or _math.isnan(d) or d == 0.0:
                return ''
            result = round_percentage_to_integer_string(n / d * 100)
            return result if result is not None else ''
        except (TypeError, ValueError, ZeroDivisionError):
            return ''

    q_act_this_tag = Tag(
        name='Q_ACT_THIS',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.act_q_this),
        tag_settings=tag_settings,
        data=True,
    )

    q_act_last_tag = Tag(
        name='Q_ACT_LAST',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.act_q_last),
        tag_settings=tag_settings,
        data=True,
    )

    norm_act_tag = Tag(
        name='NORM_ACT',
        get_value_fn=lambda obj, **kwargs: round_discharge_to_comma_separated_string(obj.act_norm),
        tag_settings=tag_settings,
        data=True,
    )

    perc_prevyear_act_tag = Tag(
        name='PERC_PREVYEAR_ACT',
        get_value_fn=lambda obj, **kwargs: _safe_perc(obj.act_q_this, obj.act_q_last),
        tag_settings=tag_settings,
        data=True,
    )

    perc_norm_act_tag = Tag(
        name='PERC_NORM_ACT',
        get_value_fn=lambda obj, **kwargs: _safe_perc(obj.act_q_this, obj.act_norm),
        tag_settings=tag_settings,
        data=True,
    )
    # endregion tags

    report_settings.report_output_path = os.getenv("ieasyreports_report_output_path")
    
    if sapphire_forecast_horizon == 'pentad':
        tag_list = [pentad_tag, forecast_tag, header_tag, river_ru_tag, punkt_ru_tag,
                model_tag, forecast_tag, dash_tag, linreg_predictor_tag,
                hydrograph_max_tag, hydrograph_min_tag, hydrograph_norm_tag,
                month_string_nom_ru_tag, month_string_gen_ru_tag, year_tag,
                day_start_pentad_tag, day_end_pentad_tag,
                delta_tag, sdivsigma_tag,
                forecast_lower_bound_tag, forecast_upper_bound_tag,
                qdanger_tag, perc_norm_tag, q_last_year_tag,
                q_act_this_tag, q_act_last_tag, norm_act_tag,
                perc_prevyear_act_tag, perc_norm_act_tag]
        report_settings.report_output_path = os.path.join(
            report_settings.report_output_path,
            "bulletins",
            "pentad",
            str(header_df['year'].values[0]))
        template_file_name=os.getenv("ieasyforecast_template_pentad_bulletin_file")

    elif sapphire_forecast_horizon == 'decad':
        tag_list = [decad_tag, forecast_tag, header_tag, river_ru_tag, punkt_ru_tag,
                model_tag, forecast_tag, dash_tag, linreg_predictor_tag,
                hydrograph_max_tag, hydrograph_min_tag, hydrograph_norm_tag,
                month_string_nom_ru_tag, month_string_gen_ru_tag, year_tag,
                day_start_decad_tag, day_end_decad_tag,
                delta_tag, sdivsigma_tag,
                forecast_lower_bound_tag, forecast_upper_bound_tag,
                qdanger_tag, perc_norm_tag, q_last_year_tag,
                q_act_this_tag, q_act_last_tag, norm_act_tag,
                perc_prevyear_act_tag, perc_norm_act_tag]

        report_settings.report_output_path = os.path.join(
            report_settings.report_output_path,
            "bulletins",
            "decad",
            str(header_df['year'].values[0]))
        template_file_name = os.getenv("ieasyforecast_template_decad_bulletin_file")

    elif sapphire_forecast_horizon == 'month':
        report_settings.report_output_path = os.path.join(
            report_settings.report_output_path,
            "bulletins", "month",
            str(header_df['year'].values[0]))
        template_file_name = os.getenv("ieasyforecast_template_month_bulletin_file")

        # Split sites into non-reservoirs (section 0) and reservoirs (sections 1 & 2)
        non_reservoirs = [s for s in bulletin_sites if 'вдхр' not in (s.punkt_name_ru or '')]
        reservoirs = [s for s in bulletin_sites if 'вдхр' in (s.punkt_name_ru or '')]
        has_quarterly = any(
            getattr(s, 'forecast_q_min_q', None) is not None
            or getattr(s, 'forecast_q_max_q', None) is not None
            for s in reservoirs
        )

        non_reservoirs = oder_sites_list_according_to_bulletin_order(non_reservoirs)
        reservoirs = oder_sites_list_according_to_bulletin_order(reservoirs)

        # Section 0: non-reservoirs with HEADER tag — monthly attributes
        sec0_tags = [
            Tag(name='BASIN_NAME', get_value_fn=lambda obj, **kwargs: obj.basin_ru,
                tag_settings=tag_settings, header=True),
            Tag(name='RIVER_NAME', get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='PUNKT_NAME', get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='NORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_LAST_YEAR', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]

        # Section 1: reservoirs monthly — same monthly attributes
        sec1_tags = [
            Tag(name='RIVER_NAME', get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='PUNKT_NAME', get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='NORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_LAST_YEAR', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]

        # Section 2: reservoirs quarterly — _q-suffixed attributes
        sec2_tags = [
            Tag(name='RIVER_NAME', get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='PUNKT_NAME', get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_min_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_max_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_min_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_max_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='NORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_norm_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_LAST_YEAR', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
        ]

        # Union of all per-section tags plus general tags for the generator's tag registry.
        # Legacy tags (VNORM, PERC_NORM, PERC_PREVYEAR) are kept in union_tags so that
        # any template or code path that still references them by name does not break.
        all_section_tags = sec0_tags + sec1_tags + sec2_tags
        seen_names: "set[str]" = set()
        union_tags = [
            fc_month_tag, fc_year_tag, fc_prevyear_tag,
            fc_month_start_tag, fc_month_end_tag,
            # Legacy tags retained for backward compatibility — not used by the monthly template
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]
        seen_names.update({"VNORM", "PERC_NORM", "PERC_PREVYEAR"})
        for t in all_section_tags:
            if t.name not in seen_names:
                union_tags.append(t)
                seen_names.add(t.name)

        year = int(header_df['year'].values[0])
        month_num = int(header_df['month_number'].values[0])
        month_name = header_df['month_str_nom_ru'].values[0]
        bulletin_file_name = f"{year}_{month_num:02}_{month_name}_monthly_forecast_bulletin.xlsx"

        report_generator = MultiSectionReportGenerator(
            tags=union_tags,
            template=template_file_name,
            templates_directory_path=os.getenv("ieasyreports_templates_directory_path"),
            reports_directory_path=report_settings.report_output_path,
            tag_settings=tag_settings,
            requires_header=True,
            tags_per_section=[sec0_tags, sec1_tags, sec2_tags],
        )
        report_generator.validate()
        report_generator.generate_report_multi(
            list_objects_per_section=[
                non_reservoirs or None,
                reservoirs or None,
                reservoirs if has_quarterly else None,
            ],
            output_filename=bulletin_file_name,
        )
        return

    elif sapphire_forecast_horizon == 'season':
        report_settings.report_output_path = os.path.join(
            report_settings.report_output_path,
            "bulletins", "season",
            str(header_df['year'].values[0]))
        template_file_name = os.getenv("ieasyforecast_template_season_bulletin_file")

        non_reservoirs = [s for s in bulletin_sites if 'вдхр' not in (s.punkt_name_ru or '')]
        reservoirs = [s for s in bulletin_sites if 'вдхр' in (s.punkt_name_ru or '')]

        non_reservoirs = oder_sites_list_according_to_bulletin_order(non_reservoirs)
        reservoirs = oder_sites_list_according_to_bulletin_order(reservoirs)

        # Section 0: non-reservoirs with HEADER tag — seasonal attributes
        sec0_tags = [
            Tag(name='BASIN_NAME', get_value_fn=lambda obj, **kwargs: obj.basin_ru,
                tag_settings=tag_settings, header=True),
            Tag(name='RIVER_NAME', get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='PUNKT_NAME', get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='NORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_LAST_YEAR', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]

        # Section 1: reservoirs — same seasonal attributes (reuse same attr names)
        sec1_tags = [
            Tag(name='RIVER_NAME', get_value_fn=lambda obj, **kwargs: obj.river_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='PUNKT_NAME', get_value_fn=lambda obj, **kwargs: obj.punkt_name_ru,
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_q_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MIN', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_min', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='V_MAX', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_v_max', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='NORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='Q_LAST_YEAR', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_prevyear_q', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]

        all_section_tags = sec0_tags + sec1_tags
        seen_names: "set[str]" = set()
        union_tags = [
            fc_month_tag, fc_year_tag, fc_prevyear_tag,
            fc_month_start_tag, fc_month_end_tag,
            # Legacy tags retained for backward compatibility — used by the non-_tj seasonal template
            Tag(name='VNORM', get_value_fn=lambda obj, **kwargs: _fmt_discharge(getattr(obj, 'forecast_vnorm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_NORM', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_norm', None)),
                tag_settings=tag_settings, data=True),
            Tag(name='PERC_PREVYEAR', get_value_fn=lambda obj, **kwargs: _fmt_percentage(getattr(obj, 'perc_prevyear', None)),
                tag_settings=tag_settings, data=True),
        ]
        seen_names.update({"VNORM", "PERC_NORM", "PERC_PREVYEAR"})
        for t in all_section_tags:
            if t.name not in seen_names:
                union_tags.append(t)
                seen_names.add(t.name)

        year = int(header_df['year'].values[0])
        month_num = int(header_df['month_number'].values[0])
        month_name = header_df['month_str_nom_ru'].values[0]
        bulletin_file_name = f"{year}_{month_num:02}_{month_name}_seasonal_forecast_bulletin.xlsx"

        report_generator = MultiSectionReportGenerator(
            tags=union_tags,
            template=template_file_name,
            templates_directory_path=os.getenv("ieasyreports_templates_directory_path"),
            reports_directory_path=report_settings.report_output_path,
            tag_settings=tag_settings,
            requires_header=True,
            tags_per_section=[sec0_tags, sec1_tags],
        )
        report_generator.validate()
        report_generator.generate_report_multi(
            list_objects_per_section=[
                non_reservoirs or None,
                reservoirs or None,
            ],
            output_filename=bulletin_file_name,
        )
        return

    # From bulletin_sites get site lists for each unique basin
    # Create a list of unique basins
    basins = [site.basin_ru for site in bulletin_sites]
    unique_basins = list(set(basins))

    # Create a list of sites for each unique basin
    sites_by_basin = {basin: [site for site in bulletin_sites if site.basin_ru == basin] for basin in unique_basins}

    # Add bulletin_sitest to sites_by_basin under basin 'all_basins'
    sites_by_basin['all_basins'] = bulletin_sites

    # Print the keys in object sites_by_basin
    print(f"DEBUG: write_to_excel: sites_by_basin keys: {sites_by_basin.keys()}")

    # Iterate over the unique basins and generate a report for each basin
    for basin in sites_by_basin.keys():
        print(f"DEBUG: write_to_excel: Generating report for basin {basin} ...")
        # Get the sites for the current basin
        sites = sites_by_basin[basin]

        # Order the sites according to the bulletin order
        sites = oder_sites_list_according_to_bulletin_order(sites)

        # Define the bulletin file name
        bulletin_file_name = f"{str(header_df['year'].values[0])}_{header_df['month_number'].values[0]:02}_{header_df['month_str_nom_ru'].values[0]}_{basin}_short_term_forecast_bulletin.xlsx"
        temp_bulletin_file_name = f"_temp_{bulletin_file_name}"

        # Generate the report
        report_generator = DefaultReportGenerator(
            tags=tag_list,
            template=template_file_name,
            templates_directory_path=os.getenv("ieasyreports_templates_directory_path"),
            reports_directory_path=report_settings.report_output_path,
            tag_settings=tag_settings,
            requires_header=True
        )

        report_generator.validate()

        report_generator.generate_report(
            list_objects=sites,
            output_filename=temp_bulletin_file_name
        )

        copy_worksheet(
            report_settings, temp_bulletin_file_name, bulletin_file_name,
            header_df, sapphire_forecast_horizon)


    # Done with the report generation

    # Note all objects that are passed to generate_report through list_obsjects
    # should be 'data' tags. 'data' tags are listed below a 'header' tag.



